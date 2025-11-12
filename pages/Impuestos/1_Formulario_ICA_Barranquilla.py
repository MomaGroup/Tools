import streamlit as st
import io
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl import Workbook
import zipfile
from datetime import datetime
from ui_utils import aplicar_css_global

# ==========================================================
# 🔐 VERIFICACIÓN DE AUTENTICACIÓN
# ==========================================================
if 'autenticado' not in st.session_state:
    st.session_state.autenticado = False

if not st.session_state.autenticado:
    st.error("🔒 Debes iniciar sesión primero")
    st.info("👈 Ve a la página principal para autenticarte")
    st.stop()

# ==========================================================
# 🎨 APLICAR ESTILOS GLOBALES
# ==========================================================
aplicar_css_global()

# ==========================================================
# 🏛 CONFIGURACIÓN DE PÁGINA
# ==========================================================
st.set_page_config(
    page_title="Formulario ICA Barranquilla",
    layout="wide",
    page_icon="🏛️"
)

# ==========================================================
# 🧭 SIDEBAR CON INFORMACIÓN DE USUARIO
# ==========================================================
with st.sidebar:
    st.markdown("---")
    st.success(f"👤 Usuario: **{st.session_state.username}**")
    if st.button("🚪 Cerrar Sesión", key="logout_ica"):
        st.session_state.autenticado = False
        st.session_state.username = None
        st.switch_page("Home.py")

# ==========================================================
# 🏛 CONFIGURACIÓN DE PÁGINA
# ==========================================================
st.set_page_config(
    page_title="Formulario ICA Barranquilla",
    layout="wide",
    page_icon="🏛️"
)

st.title("🧾 Generador de Formulario ICA - Barranquilla")
st.markdown("---")

# ==============================================================
# INTERFAZ DE CARGA DE ARCHIVOS
# ==============================================================

col1, col2 = st.columns(2)

with col1:
    st.subheader("📊 Anexos de Balance")
    uploaded_balances = st.file_uploader(
        "Sube uno o varios archivos de Anexos de balance por terceros",
        accept_multiple_files=True,
        type=["xlsx"],
        key="balances_ICA",
        help="Archivos Excel con el balance de tus empresas"
    )

with col2:
    st.subheader("📑 Datos Generales")
    uploaded_datos = st.file_uploader(
        "Sube el archivo 'Datos generales'",
        type=["xlsx"],
        key="datos_ICA",
        help="Archivo con información de tarifas y CIIU"
    )

st.markdown("---")

# ==============================================================
# BOTÓN DE PROCESAMIENTO
# ==============================================================

if st.button("⚙️ Generar Formularios ICA", type="primary", use_container_width=True):

    # Validar que existan archivos cargados
    if not uploaded_balances:
        st.error("⚠️ Por favor sube al menos un archivo de balance")
    elif not uploaded_datos:
        st.error("⚠️ Por favor sube el archivo de datos generales")
    else:
        # ==============================================================
        # PROCESAR ARCHIVO DE DATOS GENERALES
        # ==============================================================

        import openpyxl

        wb_datos = openpyxl.load_workbook(uploaded_datos)
        ws_tarifas = wb_datos["Tarifas"]

        # Crear diccionario de tarifas por NIT
        datos_tarifas = {}
        for row in ws_tarifas.iter_rows(min_row=2, values_only=True):
            if row[0]:
                nit_tabla = str(row[0]).strip()
                datos_tarifas[nit_tabla] = {
                    'ciiu': row[4] if row[4] else "",
                    'actividad': row[5] if row[5] else ""
                }

        st.success(f"✅ Datos de tarifas cargados: {len(datos_tarifas)} registros encontrados.")

    # ==========================================================
    # FUNCIÓN DE ENCABEZADO
    # ==========================================================
    def crear_encabezado(wb, nombre_empresa, nit, fecha, ciiu, actividad):
        ws = wb.active
        fill_azul = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        fill_amarillo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        font_blanca = Font(name='Arial', size=11, bold=True, color="FFFFFF")
        font_negra_bold = Font(name='Arial', size=11, bold=True, color="000000")
        alineacion_centro = Alignment(horizontal='center', vertical='center')
        borde_gris = Border(
            left=Side(style='thin', color='808080'),
            right=Side(style='thin', color='808080'),
            top=Side(style='thin', color='808080'),
            bottom=Side(style='thin', color='808080')
        )

        nit_limpio = nit.split()[0].split('-')[0].strip() if nit else ""
        nit_formateado = f"NIT {nit_limpio}"

        # Encabezado azul
        for fila in range(1, 5):
            ws.merge_cells(f'A{fila}:B{fila}')
        etiquetas = [nombre_empresa, nit_formateado, "IMPUESTO DE INDUSTRIA Y COMERCIO BARRANQUILLA", fecha]
        for i, texto in enumerate(etiquetas, 1):
            ws[f'A{i}'] = texto
            ws[f'A{i}'].fill = fill_azul
            ws[f'A{i}'].font = font_blanca
            ws[f'A{i}'].alignment = alineacion_centro

        # Bloques amarillos
        for fila, texto in enumerate([
            f"Código CIIU: {ciiu}",
            f"Actividad económica: {actividad}"
        ], start=5):
            ws.merge_cells(f'A{fila}:B{fila}')
            ws[f'A{fila}'] = texto
            ws[f'A{fila}'].fill = fill_amarillo
            ws[f'A{fila}'].font = font_negra_bold
            ws[f'A{fila}'].alignment = alineacion_centro
            for celda in ws[f'A{fila}:B{fila}'][0]:
                celda.border = borde_gris

        ws.merge_cells('A7:B7')
        ws['A7'] = ""
        for celda in ws['A7:B7'][0]:
            celda.border = borde_gris

        return ws

    # ==========================================================
    # CREAR BUFFER PARA ZIP
    # ==========================================================
    zip_buffer = io.BytesIO()
    archivos_generados = []
    
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        
        # ==========================================================
        # PROCESAMIENTO DE ARCHIVOS
        # ==========================================================
        for idx, balance_file in enumerate(uploaded_balances, 1):
            st.write("---")
            st.subheader(f"📂 Procesando archivo {idx}/{len(uploaded_balances)}: {balance_file.name}")

            try:
                wb_balance = openpyxl.load_workbook(balance_file)
                ws_balance = wb_balance.active

                nombre_empresa = ws_balance['A3'].value
                nit = ws_balance['A4'].value
                fecha = ws_balance['A5'].value

                if not nombre_empresa or not nit:
                    st.warning(f"Datos incompletos en {balance_file.name}")
                    continue

                nit_busqueda = nit.split()[0].split('-')[0].strip()
                ciiu = datos_tarifas.get(nit_busqueda, {}).get('ciiu', "")
                actividad = datos_tarifas.get(nit_busqueda, {}).get('actividad', "")

                wb_formulario = Workbook()
                ws_formulario = wb_formulario.active
                ws_formulario.title = "Formulario_ICA"

                ws_formulario = crear_encabezado(wb_formulario, nombre_empresa, nit, fecha, ciiu, actividad)
                ws_formulario.column_dimensions['A'].width = 80
                ws_formulario.column_dimensions['B'].width = 20

                # ==========================================================
                # LECTURA Y CÁLCULOS DE INGRESOS
                # ==========================================================
                df = pd.read_excel(balance_file, header=7)
                df.columns = [str(c).strip().lower() for c in df.columns]
                
                # Identificación dinámica de columnas
                mapa_columnas = {
                    'nivel': None,
                    'código cuenta contable': None,
                    'nombre cuenta contable': None,
                    'movimiento débito': None,
                    'movimiento crédito': None
                }
                
                for col in df.columns:
                    if 'nivel' in col:
                        mapa_columnas['nivel'] = col
                    elif 'código' in col and 'cuenta' in col:
                        mapa_columnas['código cuenta contable'] = col
                    elif 'nombre' in col and 'cuenta' in col:
                        mapa_columnas['nombre cuenta contable'] = col
                    elif 'débito' in col or 'debito' in col:
                        mapa_columnas['movimiento débito'] = col
                    elif 'crédito' in col or 'credito' in col:
                        mapa_columnas['movimiento crédito'] = col
                
                faltantes = [k for k, v in mapa_columnas.items() if v is None]
                if faltantes:
                    st.error(f"Faltan columnas en {balance_file.name}: {', '.join(faltantes)}")
                    continue
                
                # Normalización y filtrado
                df = df.rename(columns={v: k for k, v in mapa_columnas.items() if v})
                df = df[df['nivel'].astype(str).str.strip().str.lower() == 'auxiliar']
                df['nombre cuenta contable'] = df['nombre cuenta contable'].astype(str).str.lower()
                
                # Condiciones corregidas
                cond_ingresos = (
                    df['código cuenta contable'].astype(str).str.startswith('4') &
                    ~df['nombre cuenta contable'].str.contains('descuento|devoluci', regex=True)
                )
                
                cond_desc_dev = (
                    df['código cuenta contable'].astype(str).str.startswith('4') &
                    df['nombre cuenta contable'].str.contains('descuento|devoluci', regex=True)
                )
                
                # Aplicación de filtros
                df_ingresos = df[cond_ingresos]
                df_desc_dev = df[cond_desc_dev]
                
                # Cálculos
                total_ingresos = df_ingresos['movimiento crédito'].sum() - df_ingresos['movimiento débito'].sum()
                menos_desc_dev = df_desc_dev['movimiento débito'].sum() - df_desc_dev['movimiento crédito'].sum()

                # ==========================================================
                # ESCRITURA DE RESULTADOS EN EL FORMULARIO
                # ==========================================================
                fill_gris = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                fill_blanco = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                borde_gris = Border(
                    left=Side(style='thin', color='808080'),
                    right=Side(style='thin', color='808080'),
                    top=Side(style='thin', color='808080'),
                    bottom=Side(style='thin', color='808080')
                )
                font_arial = Font(name='Arial', size=11)
                font_arial_bold = Font(name='Arial', size=11, bold=True)
                alineacion_izq = Alignment(horizontal='left', vertical='center')

                filas = [
                    ("A8", "TOTAL INGRESOS ORDINARIOS Y EXTRAORDINARIOS DEL PERIODO EN TODO EL PAÍS", f"=MROUND({total_ingresos},1000)", fill_gris, font_arial_bold),
                    ("A9", "MENOS INGRESOS FUERA DE ESTE MUNICIPIO O DISTRITO", "=MROUND(0,1000)", fill_blanco, font_arial),
                    ("A10", "TOTAL INGRESOS ORDINARIOS Y EXTRAORDINARIOS EN ESTE MUNICIPIO", "=B8-B9", fill_gris, font_arial_bold),
                    ("A11", "MENOS INGRESOS POR DEVOLUCIONES, REBAJAS, DESCUENTOS", f"=MROUND({menos_desc_dev},1000)", fill_blanco, font_arial),
                    ("A12", "TOTAL INGRESOS GRAVABLES", "=B10-B11", fill_gris, font_arial_bold)
                ]

                for colA, texto, valor, relleno, fuente in filas:
                    fila = colA[1:]
                    ws_formulario[colA] = texto
                    ws_formulario[f"A{fila}"].font = fuente
                    ws_formulario[f"A{fila}"].fill = relleno
                    ws_formulario[f"A{fila}"].alignment = alineacion_izq
                    ws_formulario[f"B{fila}"] = valor
                    ws_formulario[f"B{fila}"].font = fuente
                    ws_formulario[f"B{fila}"].fill = relleno
                    ws_formulario[f"B{fila}"].number_format = '#,##0.00'

                for fila in ws_formulario['A8:B12']:
                    for celda in fila:
                        celda.border = borde_gris

                # ==========================================================
                # CALCULAR IMPUESTOS (ICA, Avisos, Bomberos)
                # ==========================================================
                tarifa_ica, tarifa_avisos, tarifa_bomberos = 0, 0, 0
                for row in ws_tarifas.iter_rows(min_row=2, values_only=True):
                    if row[0] and str(row[0]).strip() == nit_busqueda:
                        tarifa_ica = row[6] if row[6] else 0
                        tarifa_avisos = row[7] if row[7] else 0
                        tarifa_bomberos = row[8] if row[8] else 0
                        break

                ws_formulario['A13'] = "TOTAL IMPUESTO DE INDUSTRIA Y COMERCIO"
                ws_formulario['B13'] = f"=MROUND(B12*{tarifa_ica},1000)"
                ws_formulario['A14'] = "IMPUESTO DE AVISOS Y TABLEROS"
                ws_formulario['B14'] = f"=MROUND(B13*{tarifa_avisos},1000)"
                ws_formulario['A15'] = "SOBRETASA BOMBERIL"
                ws_formulario['B15'] = f"=MROUND(B13*{tarifa_bomberos},1000)"
                ws_formulario['A16'] = "TOTAL IMPUESTO A CARGO"
                ws_formulario['B16'] = "=B13+B14+B15"

                for fila in ws_formulario['A13:B16']:
                    for celda in fila:
                        celda.font = font_arial
                        celda.border = borde_gris
                        celda.number_format = '#,##0.00'
                
                ws_formulario['A16'].font = font_arial_bold
                ws_formulario['B16'].font = font_arial_bold
                ws_formulario['A16'].fill = fill_gris
                ws_formulario['B16'].fill = fill_gris

                # ==========================================================
                # CREAR HOJA Detalle_ICA
                # ==========================================================
                ws_detalle = wb_formulario.create_sheet("Detalle_ICA")
                alineacion_centro = Alignment(horizontal='center', vertical='center')
                alineacion_der = Alignment(horizontal='right', vertical='center')

                encabezados = ["Concepto", "Código cuenta contable", "Nombre cuenta contable", "Identificación", "Nombre tercero", "Valor"]
                for col, texto in enumerate(encabezados, 1):
                    celda = ws_detalle.cell(row=1, column=col, value=texto)
                    celda.fill = fill_gris
                    celda.font = font_arial_bold
                    celda.alignment = alineacion_centro
                    celda.border = borde_gris

                ws_detalle.column_dimensions['A'].width = 40
                ws_detalle.column_dimensions['B'].width = 18
                ws_detalle.column_dimensions['C'].width = 45
                ws_detalle.column_dimensions['D'].width = 20
                ws_detalle.column_dimensions['E'].width = 35
                ws_detalle.column_dimensions['F'].width = 18

                df_aux = df[df['nivel'].astype(str).str.strip().str.lower() == 'auxiliar']
                df_practicadas = df_aux[df_aux['código cuenta contable'].astype(str).str.startswith('2368')].copy()
                df_practicadas['valor'] = df_practicadas['movimiento crédito'] - df_practicadas['movimiento débito']

                df_sufridas = df_aux[df_aux['código cuenta contable'].astype(str).str.startswith('135518')].copy()
                df_sufridas['valor'] = df_sufridas['movimiento débito'] - df_sufridas['movimiento crédito']

                fila_actual = 1
                def escribir_detalle(df_detalle, concepto, fila_inicio):
                    fila = fila_inicio + 1
                    ws_detalle.cell(row=fila, column=6, value=0).number_format = '#,##0.00'
                    ws_detalle.cell(row=fila, column=6).border = borde_gris
                    fila += 1
                    for _, reg in df_detalle.iterrows():
                        ws_detalle.cell(row=fila, column=1, value=concepto).font = font_arial
                        ws_detalle.cell(row=fila, column=2, value=reg['código cuenta contable']).font = font_arial
                        ws_detalle.cell(row=fila, column=3, value=reg['nombre cuenta contable']).font = font_arial
                        ws_detalle.cell(row=fila, column=6, value=reg['valor']).number_format = '#,##0.00'
                        for c in range(1, 7):
                            ws_detalle.cell(row=fila, column=c).border = borde_gris
                        fila += 1
                    ws_detalle.cell(row=fila, column=5, value="Subtotal").font = font_arial_bold
                    ws_detalle.cell(row=fila, column=5).alignment = alineacion_der
                    celda_total = ws_detalle.cell(row=fila, column=6, value=f"=SUM(F{fila_inicio+1}:F{fila-1})")
                    celda_total.font = font_arial_bold
                    celda_total.number_format = '#,##0.00'
                    celda_total.border = borde_gris
                    return fila

                fila_sub_practicadas = escribir_detalle(df_practicadas, "Retenciones practicadas (2368)", fila_actual)
                fila_sub_sufridas = escribir_detalle(df_sufridas, "Retenciones sufridas (135518)", fila_sub_practicadas + 2)

                # ==========================================================
                # ENLACES Y CÁLCULOS FINALES
                # ==========================================================
                ws_formulario['A17'] = "TOTAL RETENCIONES A TÍTULO DE INDUSTRIA Y COMERCIO PRACTICADAS EN EL PERIODO"
                ws_formulario['B17'] = f"=MROUND(Detalle_ICA!F{fila_sub_practicadas},1000)"
                ws_formulario['A18'] = "(-) RETENCIONES A TÍTULO DE INDUSTRIA Y COMERCIO QUE LE PRACTICARON DURANTE EL PERIODO"
                ws_formulario['B18'] = f"=MROUND(Detalle_ICA!F{fila_sub_sufridas},1000)"
                ws_formulario['A19'] = "TOTAL RETENCIONES"
                ws_formulario['B19'] = "=B16+B17-B18"
                ws_formulario['A20'] = "SANCIONES"
                ws_formulario['B20'] = "=MROUND(0,1000)"
                ws_formulario['A21'] = "INTERESES DE MORA"
                ws_formulario['B21'] = "=MROUND(0,1000)"
                ws_formulario['A22'] = "TOTAL VALOR A PAGAR"
                ws_formulario['B22'] = "=B19+B20+B21"

                for fila in ws_formulario['A17:B22']:
                    for celda in fila:
                        celda.border = borde_gris
                        celda.font = font_arial
                        celda.number_format = '#,##0.00'
                
                # Aplicar negrita y fondo gris a filas específicas
                for fila_num in [19, 22]:
                    ws_formulario[f'A{fila_num}'].font = font_arial_bold
                    ws_formulario[f'B{fila_num}'].font = font_arial_bold
                    ws_formulario[f'A{fila_num}'].fill = fill_gris
                    ws_formulario[f'B{fila_num}'].fill = fill_gris

                # ==========================================================
                # CREAR HOJA ANEXO_BALANCE
                # ==========================================================
                ws_anexo = wb_formulario.create_sheet("Anexo_Balance")
                encabezados = [c.value for c in ws_balance[8]]
                for col, texto in enumerate(encabezados, 1):
                    celda = ws_anexo.cell(row=1, column=col, value=texto)
                    celda.fill = fill_gris
                    celda.font = font_arial_bold
                    celda.alignment = Alignment(horizontal='center', vertical='center')
                    celda.border = borde_gris

                for i, row in enumerate(ws_balance.iter_rows(min_row=9, values_only=True), start=2):
                    for j, value in enumerate(row, start=1):
                        celda = ws_anexo.cell(row=i, column=j, value=value)
                        celda.font = font_arial
                        celda.border = borde_gris
                        if isinstance(value, (int, float)):
                            celda.alignment = Alignment(horizontal='right', vertical='center')
                            celda.number_format = '#,##0.00'

                for col in ws_anexo.columns:
                    max_len = max(len(str(c.value)) if c.value else 0 for c in col)
                    ws_anexo.column_dimensions[col[0].column_letter].width = max_len + 2

                # ==========================================================
                # FORZAR RECÁLCULO AUTOMÁTICO
                # ==========================================================
                wb_formulario._calcPr = openpyxl.workbook.properties.CalcProperties(fullCalcOnLoad=True)

                # ==========================================================
                # OCULTAR CUADRÍCULAS
                # ==========================================================
                ws_formulario.sheet_view.showGridLines = False
                ws_detalle.sheet_view.showGridLines = False
                ws_anexo.sheet_view.showGridLines = False

                # ==========================================================
                # AGREGAR AL ZIP
                # ==========================================================
                nombre_archivo = f"Formulario_ICA_{nombre_empresa}.xlsx"
                buffer = io.BytesIO()
                wb_formulario.save(buffer)
                buffer.seek(0)
                
                zip_file.writestr(nombre_archivo, buffer.getvalue())
                archivos_generados.append(nombre_empresa)

                st.success(f"✅ Formulario generado correctamente para {nombre_empresa}")

            except Exception as e:
                st.error(f"❌ Error procesando {balance_file.name}: {e}")

    # ==========================================================
    # DESCARGAR ZIP AUTOMÁTICAMENTE
    # ==========================================================
    if archivos_generados:
        zip_buffer.seek(0)
        fecha_actual = datetime.now().strftime("%Y%m%d_%H%M%S")
        nombre_zip = f"Formularios_ICA_{fecha_actual}.zip"
        
        st.success(f"✅ Proceso completado. {len(archivos_generados)} formularios generados.")
        
        st.download_button(
            label=f"📦 Descargar todos los formularios ({len(archivos_generados)} archivos)",
            data=zip_buffer,
            file_name=nombre_zip,
            mime="application/zip",
            type="primary"
        )
        
        with st.expander("📋 Ver archivos generados"):
            for i, empresa in enumerate(archivos_generados, 1):
                st.write(f"{i}. {empresa}")
                

# ==========================================================
# 📖 INFORMACIÓN ADICIONAL
# ==========================================================
with st.expander("📘 Instrucciones de Uso"):
    st.markdown("""
    ### 📋 Cómo usar este generador de Formularios ICA:
    
    1. **Archivos de Balance:**  
       Sube uno o varios archivos Excel con los balances contables de tus empresas.  
       - Deben incluir la información del NIT, razón social y periodo.  
       - Estructura compatible con anexos contables estándar.
    
    2. **Datos Generales:**  
       Carga el archivo con la hoja **"Tarifas"**, que debe tener las columnas:  
       NIT, CIIU, Actividad económica y Tarifa ICA correspondiente.
    
    3. **Procesar:**  
       Haz clic en **"Generar Formularios ICA"** para iniciar el cálculo.
    
    4. **Descargar:**  
       Una vez finalizado, podrás descargar el ZIP con todos los formularios.
    """)

# ==========================================================
# 🔧 CARACTERÍSTICAS DEL FORMULARIO ICA
# ==========================================================
with st.expander("🧩 Características del Formulario ICA"):
    st.markdown("""
    ### ✨ Funcionalidades incluidas:
    
    - Cálculo automático del impuesto ICA por NIT y tarifa.  
    - Clasificación por actividad económica.  
    - Ajuste automático de valores netos y redondeo.  
    - Diseño compatible con formatos oficiales.  
    - Generación individual por empresa.  
    - Archivo ZIP con todos los formularios consolidados.  
    """)
