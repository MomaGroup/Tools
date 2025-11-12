import streamlit as st
import pandas as pd
import re
import unicodedata
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import io
import zipfile
from datetime import datetime
from ui_utils import aplicar_css_global

# ==========================================================
# 🔐 VERIFICACIÓN DE AUTENTICACIÓN
# ==========================================================
if 'autenticado' not in st.session_state:
    st.session_state.autenticado = False

if not st.session_state.autenticado:
    st.error("🔒 Debes iniciar sesión primero")
    st.info("👈 Ve a la página principal para autenticarte")
    st.stop()

# ==========================================================
# 🎨 APLICAR ESTILOS GLOBALES
# ==========================================================
aplicar_css_global()

# ==========================================================
# 🎨 CONFIGURACIÓN DE LA PÁGINA
# ==========================================================
st.set_page_config(
    page_title="Generador Formulario 2593",
    page_icon="🧮",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ==========================================================
# 🧭 SIDEBAR CON INFORMACIÓN DE USUARIO
# ==========================================================
with st.sidebar:
    st.markdown("---")
    st.success(f"👤 Usuario: **{st.session_state.username}**")
    if st.button("🚪 Cerrar Sesión", key="logout_simple"):
        st.session_state.autenticado = False
        st.session_state.username = None
        st.switch_page("Home.py")

# ==========================================================
# 🔧 FUNCIONES UTILITARIAS
# ==========================================================
def normalize_nit(s):
    s = str(s)
    m = re.match(r"(\d+)", s)
    return m.group(1) if m else ""

def normalizar_texto(s):
    return ''.join(c for c in unicodedata.normalize('NFD', str(s).lower()) if unicodedata.category(c) != 'Mn')

def redondear_mil(valor):
    return round(valor / 1000) * 1000 if pd.notna(valor) else 0

def limpiar_nombre_archivo(nombre):
    base = unicodedata.normalize('NFD', str(nombre))
    base = ''.join(c for c in base if unicodedata.category(c) != 'Mn')
    base = re.sub(r'[^A-Za-z0-9_]+', '_', base).strip('_')
    base = re.sub(r'_+', '_', base)
    return base or "Empresa_Desconocida"

# ==========================================================
# 📤 SUBIDA DE ARCHIVOS
# ==========================================================

col1, col2 = st.columns(2)

with col1:
    st.subheader("📊 Anexos de Balance")
    balance_files = st.file_uploader(
        "Sube uno o varios archivos de Anexos de balance por terceros",
        type=['xlsx'],
        accept_multiple_files=True,
        help="Uno o más archivos Excel con balances de prueba"
    )

with col2:
    st.subheader("📑Datos Generales")
    datos_generales_file = st.file_uploader(
        "Sube el archivo 'Datos generales'",
        type=['xlsx'],
        help="Archivo único con hojas: Empresas, tarifas, uvt"
    )

st.markdown("---")

# ==========================================================
# 🚀 PROCESAMIENTO
# ==========================================================
if st.button("⚙️ Generar Formularios SIMPLE", type="primary", use_container_width=True):
    
    if not datos_generales_file:
        st.error("❌ Debes subir el archivo 'Datos generales.xlsx'")
        st.stop()
    
    if not balance_files:
        st.error("❌ Debes subir al menos un archivo de Balance")
        st.stop()
    
    # Inicializar contenedor para archivos generados
    archivos_generados = {}
    
    with st.spinner("⏳ Procesando datos generales..."):
        try:
            # Cargar datos generales
            empresas = pd.read_excel(datos_generales_file, sheet_name="Empresas")
            tarifas_df = pd.read_excel(datos_generales_file, sheet_name="tarifas")
            uvt_df = pd.read_excel(datos_generales_file, sheet_name="uvt")
            
            empresas["_nit_norm"] = empresas["NIT"].astype(str).apply(normalize_nit)
            valor_uvt = float(uvt_df["valor_uvt"].iloc[0])
            
            st.success(f"✅ Datos generales cargados | UVT: ${valor_uvt:,.0f}")
        except Exception as e:
            st.error(f"❌ Error cargando datos generales: {e}")
            st.stop()
    
    # Procesar cada balance
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    for idx, balance_file in enumerate(balance_files):
        try:
            progress = (idx + 1) / len(balance_files)
            progress_bar.progress(progress)
            status_text.text(f"📊 Procesando {balance_file.name} ({idx + 1}/{len(balance_files)})")
            
            # Leer metadata
            balance_meta = pd.read_excel(balance_file, header=None)
            meta = balance_meta.head(10)
            
            nombre_empresa = str(meta.iat[2, 0]).strip() if pd.notna(meta.iat[2, 0]) else ""
            texto_nit = str(meta.iat[3, 0]).strip() if pd.notna(meta.iat[3, 0]) else ""
            m = re.match(r"(\d+)", texto_nit)
            nit_empresa = m.group(1) if m else ""
            periodo_texto = str(meta.iat[4, 0]).strip() if pd.notna(meta.iat[4, 0]) else ""
            
            # Validar empresa
            fila_empresa = empresas.loc[empresas["_nit_norm"] == nit_empresa]
            if fila_empresa.empty:
                st.warning(f"⚠️ NIT {nit_empresa} no encontrado en Datos generales - Omitiendo {balance_file.name}")
                continue
            
            grupo_valor = str(fila_empresa["Grupo"].iloc[0])
            ciiu_valor = str(fila_empresa["Código CIIU"].iloc[0])
            tarifa_ica = float(fila_empresa["Tarifa ICA"].iloc[0])
            
            # Leer balance de trabajo
            balance = pd.read_excel(balance_file, header=7)
            balance.columns = [str(c).strip() for c in balance.columns]
            balance_aux = balance[balance["Nivel"].astype(str).str.strip().str.lower() == "auxiliar"]
            
            # ==========================================================
            # 🧱 CREAR EXCEL
            # ==========================================================
            wb = Workbook()
            ws = wb.active
            ws.title = "Formato_2593"
            
            # Estilos
            fill_azul = PatternFill(start_color="1F3B63", end_color="1F3B63", fill_type="solid")
            fill_gris = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            fill_amarillo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            font_blanca = Font(color="FFFFFF", bold=True)
            font_negra = Font(color="000000", bold=True)
            align_centro = Alignment(horizontal="center", vertical="center")
            align_izquierda = Alignment(horizontal="left", vertical="center")
            align_derecha = Alignment(horizontal="right", vertical="center")
            borde_fino = Side(border_style="thin", color="000000")
            borde_lateral = Border(left=borde_fino, right=borde_fino)
            
            # PASO 1: Encabezado azul
            titulo_anticipo = "ANTICIPO SIMPLE BIMESTRAL AÑO 2025"
            encabezado = [nombre_empresa, f"NIT {nit_empresa}", titulo_anticipo, periodo_texto]
            for fila, texto in enumerate(encabezado, start=1):
                ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=3)
                celda = ws.cell(row=fila, column=1)
                celda.value = texto
                celda.fill = fill_azul
                celda.font = font_blanca
                celda.alignment = align_centro
                ws.row_dimensions[fila].height = 22
            
            # PASO 2: Grupo y CIIU
            bloque = [f"Grupo Régimen SIMPLE: {grupo_valor}", f"Código CIIU: {ciiu_valor}"]
            for i, texto in enumerate(bloque, start=5):
                ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=3)
                celda = ws.cell(row=i, column=1)
                celda.value = texto
                celda.fill = fill_gris
                celda.font = font_negra
                celda.alignment = align_centro
                ws.row_dimensions[i].height = 20
            
            # PASO 3: Ingresos principales
            def es_ingreso(row):
                cod = str(row["Código cuenta contable"])
                nombre = str(row["Nombre cuenta contable"]).lower()
                return cod.startswith("4") and not any(p in nombre for p in ["descuento", "descuentos", "devolución", "devoluciones"])
            
            def es_descuento(row):
                cod = str(row["Código cuenta contable"])
                nombre = str(row["Nombre cuenta contable"]).lower()
                return cod.startswith("4") and any(p in nombre for p in ["descuento", "descuentos"])
            
            def es_devolucion(row):
                cod = str(row["Código cuenta contable"])
                nombre = str(row["Nombre cuenta contable"]).lower()
                return cod.startswith("4") and any(p in nombre for p in ["devolución", "devoluciones"])
            
            ingresos = balance_aux[balance_aux.apply(es_ingreso, axis=1)]
            descuentos = balance_aux[balance_aux.apply(es_descuento, axis=1)]
            devoluciones = balance_aux[balance_aux.apply(es_devolucion, axis=1)]
            
            ingresos_brutos = redondear_mil(ingresos["Movimiento crédito"].sum() - ingresos["Movimiento débito"].sum())
            total_descuentos = redondear_mil(descuentos["Movimiento débito"].sum())
            total_devoluciones = redondear_mil(devoluciones["Movimiento débito"].sum())
            ingresos_con_descuentos = redondear_mil(ingresos_brutos - total_descuentos)
            ingresos_netos = redondear_mil(ingresos_brutos - total_descuentos - total_devoluciones)
            
            filas3 = {
                7: ("Ingresos brutos", ingresos_brutos),
                8: ("Total descuentos", total_descuentos),
                9: ("Total devoluciones", total_devoluciones),
                10: ("Ingresos con descuentos incluidos", ingresos_con_descuentos),
                11: ("Ingresos netos", ingresos_netos)
            }
            for fila, (texto, valor) in filas3.items():
                for col in range(1, 4):
                    celda = ws.cell(row=fila, column=col)
                    celda.fill = fill_amarillo
                    celda.font = font_negra
                    celda.alignment = align_derecha
                    celda.border = Border(bottom=borde_fino)
                ws.cell(row=fila, column=2).value = texto
                celda_valor = ws.cell(row=fila, column=3)
                celda_valor.value = valor
                celda_valor.number_format = '_-* #,##0_-;-* #,##0_-;_-* "-"??_-;_-@_-'
            
            # PASO 4: Ingresos bimestrales
            ingresos_brutos_bimestrales = ingresos_con_descuentos
            ingresos_no_constitutivos = 0
            total_ingresos_netos_bimestrales = ingresos_brutos_bimestrales - ingresos_no_constitutivos
            
            filas4 = {
                12: ("Ingresos brutos bimestrales", ingresos_brutos_bimestrales),
                13: ("Ingresos no constitutivos de renta", ingresos_no_constitutivos),
                14: ("Total ingresos netos bimestrales", total_ingresos_netos_bimestrales)
            }
            for fila, (texto, valor) in filas4.items():
                for col in range(1, 4):
                    celda = ws.cell(row=fila, column=col)
                    celda.fill = PatternFill(fill_type=None)
                    celda.border = borde_lateral
                    celda.font = Font(color="000000", bold=(texto == "Total ingresos netos bimestrales"))
                    celda.alignment = align_izquierda if col == 2 else align_derecha
                ws.cell(row=fila, column=2).value = texto
                celda_valor = ws.cell(row=fila, column=3)
                celda_valor.value = valor
                celda_valor.number_format = '_-* #,##0_-;-* #,##0_-;_-* "-"??_-;_-@_-'
            
            # PASO 5: Anticipo SIMPLE
            ingreso_uvt = ingresos_con_descuentos / valor_uvt
            tarifa_fila = tarifas_df[
                (tarifas_df["grupo"].str.lower() == grupo_valor.lower()) &
                (tarifas_df["uvt_min"] <= ingreso_uvt) &
                (tarifas_df["uvt_max"] >= ingreso_uvt)
            ]
            tarifa_simple = float(tarifa_fila["tarifa"].iloc[0]) if not tarifa_fila.empty else 0
            
            anticipo_simple = redondear_mil(ingresos_con_descuentos * tarifa_simple)
            ica_bimestre = redondear_mil(ingresos_netos * tarifa_ica)
            anticipo_nacional = redondear_mil(anticipo_simple - ica_bimestre)
            
            filas5 = {
                15: ("Anticipo SIMPLE", anticipo_simple),
                16: ("ICA consolidado liquidado durante el bimestre", ica_bimestre),
                17: ("Valor del anticipo componente SIMPLE nacional", anticipo_nacional)
            }
            for fila, (texto, valor) in filas5.items():
                for col in range(1, 4):
                    celda = ws.cell(row=fila, column=col)
                    celda.fill = PatternFill(fill_type=None)
                    celda.border = borde_lateral
                    celda.font = Font(color="000000", bold=(texto == "Valor del anticipo componente SIMPLE nacional"))
                    celda.alignment = align_izquierda if col == 2 else align_derecha
                ws.cell(row=fila, column=2).value = texto
                celda_valor = ws.cell(row=fila, column=3)
                celda_valor.value = valor
                celda_valor.number_format = '_-* #,##0_-;-* #,##0_-;_-* "-"??_-;_-@_-'
            
            # PASO 6: Aportes pensiones
            def es_aporte_pension(row):
                cod = str(row["Código cuenta contable"]).strip()
                nivel = str(row["Nivel"]).strip().lower()
                nombre = normalizar_texto(str(row["Nombre cuenta contable"]))
                return (nivel == "auxiliar"
                        and (cod.startswith("5") or cod.startswith("6"))
                        and ("aporte" in nombre and "pension" in nombre))
            
            aportes = balance_aux[balance_aux.apply(es_aporte_pension, axis=1)]
            aporte_valor = redondear_mil(aportes["Movimiento débito"].sum() - aportes["Movimiento crédito"].sum())
            
            filas6 = {
                18: ("Aportes al Sistema General de Pensiones a cargo del empleador, pagados en el bimestre", aporte_valor),
                19: ("Excedente de aportes al Sistema General de Pensiones a cargo del empleador no solicitado como descuento en el bimestre anterior", 0),
                20: ("Excedentes de aportes del bimestre anterior y aportes durante el bimestre al Sistema General de Pensiones a cargo del empleador (Limitados)", "=MIN(C18+C19,C17)")
            }
            for fila, (texto, valor) in filas6.items():
                for col in range(1, 4):
                    celda = ws.cell(row=fila, column=col)
                    celda.fill = PatternFill(fill_type=None)
                    celda.border = borde_lateral
                    celda.font = Font(color="000000", bold=(fila == 20))
                    celda.alignment = align_izquierda if col == 2 else align_derecha
                ws.cell(row=fila, column=2).value = texto
                celda_valor = ws.cell(row=fila, column=3)
                celda_valor.value = valor
                celda_valor.number_format = '_-* #,##0_-;-* #,##0_-;_-* "-"??_-;_-@_-'
            
            # PASO 7: Anticipo neto y saldos
            filas7 = {
                21: ("Exceso de aportes al Sistema General de Pensiones a cargo del empleador", "=C18+C19-C20"),
                22: ("Anticipo neto impuesto unificado SIMPLE", "=C17-C20"),
                23: ("Retenciones y autorretenciones a título de renta practicadas antes de pertenecer al régimen SIMPLE", "=0"),
                24: ("Excedente anticipo impuesto SIMPLE del bimestre anterior", "=0"),
                25: ("Saldo a favor por impuesto SIMPLE declaración año anterior", "=0"),
                26: ("Saldo anticipo SIMPLE", "=C22-C23-C24-C25"),
                27: ("Intereses causados por anticipo SIMPLE (aplicación art. 804 E.T.)", "=0"),
                28: ("Pagos anteriores anticipo SIMPLE por este periodo", "=0"),
                29: ("Total a pagar anticipo impuesto SIMPLE", "=C22+C27-C23-C24-C25-C28"),
                30: ("Excedente anticipo impuesto SIMPLE", "=IF((C23+C24+C25+C28-C22-C27)>0,C23+C24+C25+C28-C22-C27,0)"),
            }
            for fila, (texto, formula) in filas7.items():
                for col in range(1, 4):
                    celda = ws.cell(row=fila, column=col)
                    celda.fill = PatternFill(fill_type=None)
                    if fila in [22, 30]:
                        celda.border = Border(left=borde_fino, right=borde_fino, bottom=borde_fino)
                    else:
                        celda.border = borde_lateral
                    if texto in [
                        "Exceso de aportes al Sistema General de Pensiones a cargo del empleador",
                        "Anticipo neto impuesto unificado SIMPLE",
                        "Saldo anticipo SIMPLE",
                        "Total a pagar anticipo impuesto SIMPLE",
                        "Excedente anticipo impuesto SIMPLE",
                    ]:
                        celda.font = Font(color="000000", bold=True)
                    else:
                        celda.font = Font(color="000000", bold=False)
                    celda.alignment = align_izquierda if col == 2 else align_derecha
                ws.cell(row=fila, column=2).value = texto
                celda_valor = ws.cell(row=fila, column=3)
                celda_valor.value = formula
                celda_valor.number_format = '_-* #,##0_-;-* #,##0_-;_-* "-"??_-;_-@_-'
            
            # PASO 8: Liquidación IVA
            def es_iva_generado(row):
                cod = str(row["Código cuenta contable"]).strip()
                nombre = str(row["Nombre cuenta contable"]).lower()
                return (
                    cod.startswith("2408")
                    and any(p in nombre for p in [
                        "iva generado", "iva generado en ventas", "iva generado por servicios",
                        "iva ventas nacionales", "iva ventas gravadas",
                        "iva por operaciones gravadas", "iva en operaciones gravadas",
                        "iva generado 19", "iva generado 5",
                        "iva generado y por pagar", "iva causado ventas",
                        "iva causado por ventas", "iva causado en operaciones gravadas",
                        "iva generado operaciones gravadas", "iva generado ventas 19",
                        "iva generado servicios 19", "iva generado ventas gravadas 19",
                        "iva causado ventas nacionales", "iva causado servicios",
                        "iva generado por operaciones", "iva generado por ventas nacionales",
                        "iva generado por operaciones gravadas", "iva generado por ventas 19",
                        "iva devolucion en compras", "iva dev compras",
                        "iva devolucion en compras 19", "iva devolucion en compras 5",
                        "iva devolucion en servicios", "iva devolucion en servicios 19",
                        "iva devolucion en servicios 5", "devolucion descontable por servicios",
                        "devolucion descontable por servicios 19", "devolucion descontable por servicios 5"
                    ])
                )
            
            def es_iva_descontable(row):
                cod = str(row["Código cuenta contable"]).strip()
                nombre = str(row["Nombre cuenta contable"]).lower()
                return (
                    cod.startswith("2408")
                    and any(p in nombre for p in [
                        "iva descontable", "iva descontable en compras", "iva descontable por compras",
                        "iva descontable por servicios", "descontable por servicios",
                        "iva descontable compras nacionales", "iva descontable por compras 19%",
                        "iva descontable compras gravadas", "iva compras", "iva compras nacionales",
                        "iva compras gravadas", "iva servicios", "iva servicios 19",
                        "iva compras 19", "iva descontable 19", "iva descontable importaciones",
                        "descontable por devoluciones 19%", "descontable por devoluciones servicios",
                        "descontable por devoluciones", "iva descontable por compras 5%"
                    ])
                )
            
            def calcular_retenciones_iva(balance_aux):
                def _norm(s):
                    s = str(s).lower()
                    s = ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
                    return s
                
                debitos = balance_aux[
                    (balance_aux["Código cuenta contable"].astype(str).str.startswith("135517")) &
                    (balance_aux["Nivel"].astype(str).str.lower().str.strip() == "auxiliar") &
                    (balance_aux["Nombre cuenta contable"].apply(lambda x: "impuesto a las ventas retenido" in _norm(x)))
                ]
                creditos = balance_aux[
                    (balance_aux["Código cuenta contable"].astype(str).str.startswith("135517")) &
                    (balance_aux["Nivel"].astype(str).str.lower().str.strip() == "auxiliar") &
                    (balance_aux["Nombre cuenta contable"].apply(lambda x: "devolucion impuesto a las ventas retenido" in _norm(x)))
                ]
                total_debito = debitos["Movimiento débito"].sum()
                total_credito = creditos["Movimiento crédito"].sum()
                return redondear_mil(total_debito - total_credito)
            
            iva_generado = redondear_mil(balance_aux[balance_aux.apply(es_iva_generado, axis=1)]["Movimiento crédito"].sum())
            iva_descontable = redondear_mil(balance_aux[balance_aux.apply(es_iva_descontable, axis=1)]["Movimiento débito"].sum())
            retenciones_iva = calcular_retenciones_iva(balance_aux)
            
            filas8 = {
                31: ("Impuesto sobre las ventas generado en el bimestre en operaciones gravadas", iva_generado),
                32: ("Impuestos descontables del bimestre", iva_descontable),
                33: ("Saldo a favor declaración impuesto sobre las ventas año gravable anterior", 0),
                34: ("Excedente IVA bimestre anterior", 0),
                35: ("Retenciones por IVA que le practicaron en el bimestre", retenciones_iva),
                36: ("Saldo impuesto sobre las ventas", "=IF((C31-C32-C33-C34-C35)>0,C31-C32-C33-C34-C35,0)"),
                37: ("Intereses causados por impuesto sobre las ventas (aplicación art. 804 E.T.)", 0),
                38: ("Pagos anteriores impuesto sobre las ventas por este período", 0),
                39: ("Total a pagar IVA bimestral", "=IF((C31+C37-C32-C33-C34-C35-C38)>0,C31+C37-C32-C33-C34-C35-C38,0)"),
                40: ("Excedente IVA bimestral", "=IF((C32+C33+C34+C35+C38-C31-C37)>0,C32+C33+C34+C35+C38-C31-C37,0)"),
            }
            for fila, (texto, valor) in filas8.items():
                for col in range(1, 4):
                    celda = ws.cell(row=fila, column=col)
                    celda.fill = PatternFill(fill_type=None)
                    if fila == 40:
                        celda.border = Border(left=borde_fino, right=borde_fino, bottom=borde_fino)
                    else:
                        celda.border = borde_lateral
                    celda.font = Font(color="000000", bold=(fila in [36, 39, 40]))
                    celda.alignment = align_izquierda if col == 2 else align_derecha
                ws.cell(row=fila, column=2).value = texto
                celda_valor = ws.cell(row=fila, column=3)
                celda_valor.value = valor
                celda_valor.number_format = '_-* #,##0_-;-* #,##0_-;_-* "-"??_-;_-@_-'

            # ==========================================================
            # 🟩 PASO 9: Valores a pagar
            # ==========================================================
            filas9 = {
                41: ("ICA municipios y distritos", "=C16"),
                42: ("Anticipo SIMPLE", "=C29"),
                43: ("Impuesto sobre las ventas", "=C39"),
                44: ("Total a pagar", "=C41+C42+C43"),
            }
            for fila, (texto, formula) in filas9.items():
                for col in range(1, 4):
                    celda = ws.cell(row=fila, column=col)
                    celda.fill = PatternFill(fill_type=None)
                    if fila in [43, 44]:
                        celda.border = Border(left=borde_fino, right=borde_fino, bottom=borde_fino)
                    else:
                        celda.border = borde_lateral
                    celda.font = Font(color="000000", bold=(fila == 44))
                    celda.alignment = align_izquierda if col == 2 else align_derecha
                ws.cell(row=fila, column=2).value = texto
                celda_valor = ws.cell(row=fila, column=3)
                celda_valor.value = formula
                celda_valor.number_format = '_-* #,##0_-;-* #,##0_-;_-* "-"??_-;_-@_-'
            
            # ==========================================================
            # 🔷 PASO 10: Etiquetas verticales de sección (columna A)
            # ==========================================================
            bloques_columna_A = {
                (12, 22): "Liquidación anticipo bimestral componente SIMPLE nacional",
                (23, 30): "Determinación del valor a pagar o excedente anticipo SIMPLE",
                (31, 40): "Liquidación impuesto sobre las ventas",
                (41, 43): "Valores a pagar",
            }
            for (inicio, fin), texto in bloques_columna_A.items():
                ws.merge_cells(start_row=inicio, start_column=1, end_row=fin, end_column=1)
                celda = ws.cell(row=inicio, column=1)
                celda.value = texto
                celda.font = Font(color="000000", bold=True, size=10)
                celda.alignment = Alignment(horizontal="center", vertical="center", text_rotation=90, wrap_text=True)
                borde_completo = Border(left=borde_fino, right=borde_fino, top=borde_fino, bottom=borde_fino)
                for fila in range(inicio, fin + 1):
                    ws.cell(row=fila, column=1).border = borde_completo
            
            # ==========================================================
            # 🟪 PASO 11: Agregar hoja con BALANCE ORIGINAL
            # ==========================================================
            ws_balance = wb.create_sheet(title="Balance_Original")
            
            # Detectar columnas contables
            columnas_contables = [
                col for col in balance.columns
                if any(palabra in col.lower() for palabra in ["débito", "debito", "crédito", "credito", "saldo", "valor", "movimiento"])
            ]
            
            # Copiar el DataFrame original completo
            for r_idx, row in enumerate(dataframe_to_rows(balance, index=False, header=True), start=1):
                for c_idx, value in enumerate(row, start=1):
                    celda = ws_balance.cell(row=r_idx, column=c_idx, value=value)
                    encabezado = balance.columns[c_idx - 1] if c_idx - 1 < len(balance.columns) else ""
                    
                    # Encabezados en negrita y centrados
                    if r_idx == 1:
                        celda.font = Font(bold=True)
                        celda.alignment = Alignment(horizontal="center")
                    
                    # Formato contable solo para columnas alineadas a la derecha
                    elif encabezado in columnas_contables and isinstance(value, (int, float)):
                        celda.number_format = '_-* #,##0_-;-* #,##0_-;_-* "-"??_-;_-@_-'
                        celda.alignment = Alignment(horizontal="right")
            
            # Ajustar automáticamente el ancho de columnas
            for column_cells in ws_balance.columns:
                max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
                adjusted_width = min(max_length + 2, 60)
                ws_balance.column_dimensions[column_cells[0].column_letter].width = adjusted_width
            
            # ==========================================================
            # 📏 ANCHO FIJO DE COLUMNAS EN "Formato_2593"
            # ==========================================================
            ws.column_dimensions["A"].width = 8
            ws.column_dimensions["B"].width = 127.86
            ws.column_dimensions["C"].width = 12.57
            
            # ==========================================================
            # 💾 Guardar archivo en memoria
            # ==========================================================
            nombre_limpio = limpiar_nombre_archivo(nombre_empresa if nombre_empresa else "Empresa_Desconocida")
            output_filename = f"Formato_2593_{nombre_limpio}.xlsx"
            
            ws.sheet_view.showGridLines = False
            
            # Guardar en bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            # Almacenar en diccionario
            archivos_generados[output_filename] = output.read()
            
            st.success(f"✅ {nombre_empresa} | NIT: {nit_empresa} | Total a pagar: ${redondear_mil(anticipo_nacional + ica_bimestre):,.0f}")
            
        except Exception as e:
            st.error(f"❌ Error en {balance_file.name}: {str(e)}")
            continue
    
    progress_bar.progress(1.0)
    status_text.text("✅ Procesamiento completado")

    # ==========================================================
    # 📦 SECCIÓN DE DESCARGAS
    # ==========================================================
    if archivos_generados:
        st.header("📥 2. Descargar Formularios")
        
        st.success(f"🎉 Se generaron {len(archivos_generados)} formulario(s) exitosamente")
        
        # Descargas individuales
        st.subheader("Descarga Individual")
        cols = st.columns(min(len(archivos_generados), 3))
        for idx, (filename, file_bytes) in enumerate(archivos_generados.items()):
            with cols[idx % 3]:
                st.download_button(
                    label=f"📄 {filename}",
                    data=file_bytes,
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
        
        # Descarga en ZIP
        st.subheader("Descarga Masiva")
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for filename, file_bytes in archivos_generados.items():
                zip_file.writestr(filename, file_bytes)
        
        zip_buffer.seek(0)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        st.download_button(
            label="📦 Descargar Todos los Formularios (ZIP)",
            data=zip_buffer,
            file_name=f"Formularios_2593_{timestamp}.zip",
            mime="application/zip",
            use_container_width=True,
            type="primary"
        )
        
        # Resumen final
        with st.expander("📊 Ver Resumen de Archivos Generados"):
            resumen_df = pd.DataFrame([
                {"Archivo": filename, "Tamaño (KB)": f"{len(file_bytes) / 1024:.1f}"}
                for filename, file_bytes in archivos_generados.items()
            ])
            st.dataframe(resumen_df, use_container_width=True)
    
    else:
        st.warning("⚠️ No se generaron formularios. Verifica los datos de entrada.")

# ==========================================================
# 🧩 INFORMACIÓN ADICIONAL
# ==========================================================

# 📘 INSTRUCCIONES DE USO
with st.expander("📘 Instrucciones de Uso"):
    st.markdown("""
    ### 🧾 Cómo usar este generador de Formularios 2593 - Régimen SIMPLE:

    1. **📊 Archivos de Balance:**
       - Sube uno o varios archivos Excel con los balances contables de tus empresas.
       - Deben incluir la información de **nombre de la empresa, NIT, razón social y periodo**.
       - Estructura compatible con anexos contables estándar exportados de tu software contable.
    
    <br>
    
    2. **📑 Datos Generales:**
       - Carga el archivo **`Datos generales.xlsx`**, que debe contener las hojas:
         - **Empresas:** con columnas `NIT`, `Ultimo digito del NIT`, `Razon social`, `Ciudad`, `Grupo`, `Código CIIU`, `Tarifa ICA`.
         - **Tarifas:** con los rangos UVT y la tarifa SIMPLE correspondiente por grupo.
         - **ICA por ciudad:** con las tarifas ICA relacionadas por ciudad.
         - **UVT:** con el valor de la UVT vigente.
       - Asegúrate de que los NIT coincidan entre los balances y el archivo de datos generales.
    
    <br>
    
    3. **⚙️ Procesar Formularios:**
       - Haz clic en **“Generar Formularios SIMPLE”** para iniciar el cálculo.
    
    <br>
    
    4. **📥 Descargar Resultados:**
       - Una vez finalizado, podrás descargar:
         - Archivos **individuales** (`Formato_2593_Empresa.xlsx`).
         - Un **ZIP consolidado** con todos los formularios generados.
    
    <br>
    
    5. **💡 Consejos:**
       - No modifiques los nombres de las hojas del archivo “Datos generales”.
       - Si una empresa no genera formulario, revisa su NIT o estructura del balance.
       - Cada archivo generado incluye una hoja adicional llamada **Balance_Original**.
    """, unsafe_allow_html=True)

# ==========================================================
# ⚙️ CARACTERÍSTICAS DEL FORMULARIO 2593
# ==========================================================
with st.expander("🧩 Características del Formulario 2593"):
    st.markdown("""
    ### 🧮 Secciones procesadas:
    - Liquidación anticipo bimestral componente SIMPLE nacional
    - Determinación del valor a pagar o excedente anticipo SIMPLE
    - Liquidación impuesto sobre las ventas
    - Conceptos a pagar
    """)
