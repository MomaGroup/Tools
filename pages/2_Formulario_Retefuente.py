import streamlit as st
import pandas as pd
import re
import unicodedata
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from copy import copy
import openpyxl
from io import BytesIO
import traceback
import zipfile
import io
from ui_utils import aplicar_css_global

# ==========================================================
# 🔐 VERIFICACIÓN DE AUTENTICACIÓN
# ==========================================================
if 'autenticado' not in st.session_state:
    st.session_state.autenticado = False

if not st.session_state.autenticado:
    st.error("🔒 Debes iniciar sesión primero")
    st.info("👈 Ve a la página principal para autenticarte")
    st.stop()

# ==========================================================
# 🎨 APLICAR ESTILOS GLOBALES
# ==========================================================
aplicar_css_global()

# ==========================================================
# 🏛 CONFIGURACIÓN DE PÁGINA
# ==========================================================
st.set_page_config(
    page_title="Formulario 350 - Retefuente",
    layout="wide",
    page_icon="🧾"
)

st.title("🧾 Generador de Formulario Retefuente")
st.markdown("---")

# ==========================================================
# 🧭 SIDEBAR CON INFORMACIÓN DE USUARIO
# ==========================================================
with st.sidebar:
    st.markdown("---")
    st.success(f"👤 Usuario: **{st.session_state.username}**")
    if st.button("🚪 Cerrar Sesión", key="logout_retefuente"):
        st.session_state.autenticado = False
        st.session_state.username = None
        st.switch_page("Home.py")

# ==========================================================
# 🔧 FUNCIONES AUXILIARES
# ==========================================================
def norm_txt(s: str) -> str:
    """Normaliza texto (minúsculas, sin tildes, sin espacios dobles)."""
    if pd.isna(s):
        return ""
    s = str(s).strip().lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return s

def contiene_alguna(texto, patrones):
    """✅ Verifica si el texto contiene todas las palabras de algún patrón."""
    t = norm_txt(texto)
    for patron in patrones:
        subpalabras = patron.split()
        if all(sp in t for sp in subpalabras):
            return True
    return False

def es_auxiliar(fila) -> bool:
    """Evalúa si la cuenta contable es auxiliar."""
    nivel = norm_txt(fila.get("Nivel", ""))
    trx = norm_txt(fila.get("Transaccional", ""))
    return ("aux" in nivel) or (trx in {"si", "sí", "yes", "true", "1"})

# ==========================================================
# 🎨 ESTILOS
# ==========================================================
azul = PatternFill(start_color="1f497d", end_color="1f497d", fill_type="solid")
gris = PatternFill(start_color="d9d9d9", end_color="d9d9d9", fill_type="solid")
amarillo = PatternFill(start_color="ffff00", end_color="ffff00", fill_type="solid")
gris_subtotal = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

blanco_negrita = Font(color="FFFFFF", bold=True, size=11, name="Arial")
negro_negrita = Font(color="000000", bold=True, size=10, name="Arial")
centrado = Alignment(horizontal="center", vertical="center", wrap_text=True)
borde_todo = Border(
    left=Side(style="thin", color="808080"),
    right=Side(style="thin", color="808080"),
    top=Side(style="thin", color="808080"),
    bottom=Side(style="thin", color="808080")
)

# ==========================================================
# 📂 INTERFAZ DE CARGA DE ARCHIVOS
# ==========================================================
col1, col2 = st.columns(2)

with col1:
    st.subheader("📊 Anexos de Balance")
    uploaded_balances = st.file_uploader(
        "Sube uno o varios archivos de Anexos de balance por terceros",
        type=["xlsx"],
        accept_multiple_files=True,
        key="balances_350",
        help="Archivos Excel con el balance de tus empresas"
    )

with col2:
    st.subheader("📑 Datos Generales")
    uploaded_datos = st.file_uploader(
        "Sube el archivo 'Datos generales'",
        type=["xlsx"],
        key="datos_350",
        help="Archivo con información de tarifas y CIIU"
    )

st.markdown("---")

# ==========================================================
# 🚀 BOTÓN DE PROCESAMIENTO
# ==========================================================
if st.button("⚙️ Generar Formularios 350", type="primary", use_container_width=True):
    if not uploaded_balances:
        st.error("⚠️ Por favor sube al menos un archivo de balance")
    elif not uploaded_datos:
        st.error("⚠️ Por favor sube el archivo de datos generales")
    else:
        # ==========================================================
        # 📊 CARGAR DATOS GENERALES UNA SOLA VEZ
        # ==========================================================
        datos_generales_cache = None
        try:
            datos_generales_cache = pd.read_excel(uploaded_datos, sheet_name=None)
            st.success(f"✅ Datos generales cargados: {list(datos_generales_cache.keys())}")
        except Exception as e:
            st.error(f"⚠️ Error al cargar datos generales: {e}")

        # ==========================================================
        # 🔄 PROCESAR CADA BALANCE
        # ==========================================================
        archivos_procesados = 0
        archivos_totales = len(uploaded_balances)

        progress_bar = st.progress(0)
        status_text = st.empty()

        archivos_generados = []

        for idx, balance_file in enumerate(uploaded_balances, start=1):
            status_text.text(f"📘 Procesando archivo {idx}/{archivos_totales}: {balance_file.name}")

            try:
                # ==========================================================
                # 🆕 CREAR NUEVO WORKBOOK
                # ==========================================================
                wb = Workbook()
                ws = wb.active
                ws.title = "Detalle_350"

                # ==========================================================
                # 📄 LEER BALANCE ESPECÍFICO
                # ==========================================================
                raw = pd.read_excel(balance_file, header=7, dtype={"Identificación": str})

                # Verificación columnas
                cols_esperadas = [
                    "Nivel", "Transaccional", "Código cuenta contable", "Nombre cuenta contable",
                    "Identificación", "Sucursal", "Nombre tercero",
                    "Saldo inicial", "Movimiento débito", "Movimiento crédito", "Saldo final"
                ]
                faltantes = [c for c in cols_esperadas if c not in raw.columns]
                if faltantes:
                    raise ValueError(f"Faltan columnas: {faltantes}")

                # ==========================================================
                # 📋 EXTRAER METADATOS
                # ==========================================================
                meta = pd.read_excel(balance_file, header=None).head(10)
                nombre_empresa = str(meta.iat[2, 0]).strip() if pd.notna(meta.iat[2, 0]) else f"Empresa_{idx}"
                texto_nit = str(meta.iat[3, 0]).strip() if pd.notna(meta.iat[3, 0]) else ""
                periodo_texto = str(meta.iat[4, 0]).strip() if pd.notna(meta.iat[4, 0]) else ""
                m = re.search(r"(\d+)", texto_nit)
                nit_numero = m.group(1) if m else str(idx)
                nit_empresa = f"NIT {nit_numero}" if nit_numero else ""

                # ==========================================================
                # 🧾 BUSCAR DATOS EN HOJA "Tarifas"
                # ==========================================================
                ciiu, actividad, tarifa = "", "", ""
                try:
                    if datos_generales_cache and "Tarifas" in datos_generales_cache:
                        tarifas_df = datos_generales_cache["Tarifas"]
                        tarifas_df.columns = [norm_txt(c) for c in tarifas_df.columns]

                        col_nit = next((c for c in tarifas_df.columns if "nit" in c), None)
                        col_ciiu = next((c for c in tarifas_df.columns if "ciiu" in c), None)
                        col_actividad = next((c for c in tarifas_df.columns if "actividad" in c), None)
                        col_tarifa = next((c for c in tarifas_df.columns if "tarifa" in c), None)

                        if all([col_nit, col_ciiu, col_actividad, col_tarifa]):
                            filtro = tarifas_df[col_nit].astype(str).str.contains(nit_numero, na=False)
                            fila = tarifas_df[filtro]
                            if not fila.empty:
                                ciiu = str(fila.iloc[0][col_ciiu]).strip()
                                actividad = str(fila.iloc[0][col_actividad]).strip()
                                tarifa = str(fila.iloc[0][col_tarifa]).strip()
                except Exception as e:
                    st.warning(f"⚠️ Error al procesar tarifas para {nombre_empresa}: {e}")

                # ==========================================================
                # 📝 CREAR ENCABEZADO DEL EXCEL
                # ==========================================================
                encabezado = [nombre_empresa, nit_empresa, "DECLARACIÓN DE RETENCIÓN EN LA FUENTE", periodo_texto]
                for fila, texto in enumerate(encabezado, start=1):
                    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=9)
                    c = ws.cell(row=fila, column=1, value=texto)
                    c.fill = azul
                    c.font = blanco_negrita
                    c.alignment = centrado

                # Bloque gris
                bloque = [
                    f"Código CIIU: {ciiu if ciiu else '-'}",
                    f"Actividad económica: {actividad if actividad else '-'}",
                    f"Tarifa: {tarifa if tarifa else '-'}"
                ]
                for i, texto in enumerate(bloque, start=5):
                    ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=9)
                    c = ws.cell(row=i, column=1, value=texto)
                    c.fill = gris
                    c.font = negro_negrita
                    c.alignment = centrado
                    for col in range(1, 10):
                        ws.cell(row=i, column=col).border = borde_todo

                # Encabezado amarillo
                ws.merge_cells("A8:A9"); ws["A8"] = "Concepto"
                ws.merge_cells("B8:B9"); ws["B8"] = "Código cuenta contable"
                ws.merge_cells("C8:C9"); ws["C8"] = "Nombre cuenta contable"
                ws.merge_cells("D8:D9"); ws["D8"] = "Identificación"
                ws.merge_cells("E8:E9"); ws["E8"] = "Nombre tercero"
                ws.merge_cells("F8:G8"); ws["F8"] = "A personas jurídicas"
                ws.merge_cells("H8:I8"); ws["H8"] = "A personas naturales"
                ws["F9"] = "Base sujeta a retención"
                ws["G9"] = "Retenciones a título de renta"
                ws["H9"] = "Base sujeta a retención"
                ws["I9"] = "Retenciones a título de renta"

                for fila in (8, 9):
                    for col in range(1, 10):
                        c = ws.cell(row=fila, column=col)
                        c.fill = amarillo
                        c.font = negro_negrita
                        c.alignment = centrado
                        c.border = borde_todo

                # ==========================================================
                # 🔹 FUNCIÓN PARA GENERAR CONCEPTOS
                # ==========================================================
                def generar_concepto_multitarifa(nombre_concepto, palabras_incluir, fila_out, tasas):
                    df = raw.copy()
                    mask_aux = df.apply(es_auxiliar, axis=1)
                    mask_cod = df["Código cuenta contable"].astype(str).str.startswith("2365")
                    mask_nombre = df["Nombre cuenta contable"].astype(str).apply(lambda x: contiene_alguna(x, palabras_incluir))
                    mask_excluir = df["Nombre tercero"].astype(str).apply(lambda x: contiene_alguna(x, ["dian", "direccion", "impuesto", "aduana"]))
                    filtrado = df[mask_aux & mask_cod & mask_nombre & (~mask_excluir)].copy()

                    filtrado["__retencion__"] = (
                        pd.to_numeric(filtrado["Movimiento crédito"], errors="coerce").fillna(0)
                        - pd.to_numeric(filtrado["Movimiento débito"], errors="coerce").fillna(0)
                    )

                    if filtrado.empty:
                        ws.cell(row=fila_out, column=1, value=nombre_concepto).font = negro_negrita
                        for col in range(6, 10):
                            c = ws.cell(row=fila_out, column=col, value=0)
                            c.number_format = '#,##0.00'; c.border = borde_todo
                        fila_inicio = fila_out
                        fila_fin = fila_out
                        fila_out += 1

                        ws.cell(row=fila_out, column=1, value=f"Subtotal {nombre_concepto.lower()}").font = negro_negrita
                        for col in range(6, 10):
                            c = ws.cell(row=fila_out, column=col, value=f"=SUM({get_column_letter(col)}{fila_inicio}:{get_column_letter(col)}{fila_fin})")
                            c.number_format = '#,##0.00'
                            c.font = negro_negrita
                            c.border = borde_todo
                        for col in range(1, 10):
                            c = ws.cell(row=fila_out, column=col)
                            c.fill = gris_subtotal
                            c.font = negro_negrita
                            c.border = borde_todo
                        fila_out += 1
                        return fila_out

                    patrones_pj = [
                        r"\bsas\b",
                        r"\bs\.a\.\b",
                        r"\bltda\b",
                        r"empresa",
                        r"sociedad",
                        r"\bsc\b",
                        r"\bsca\b",
                        r"\bcompania\b",
                        r"\bcomandita\b",
                        r"fundacion",
                        r"universidad",
                        r"\bcia\b",
                        r"\bltda\b",
                        r"\blimitada\b",
                        r"\bresponsabilidad limitada\b",
                        r"\bsimple y por acciones\b",
                        r"\bs\.?a\.?\b",
                        r"\bs\.?a\.?s\.?\b",
                        r"\bltda\.?\b",
                        r"centro",
                        r"instituto",
                        r"\basociaci[oó]n\b",
                        r"\bconsorcio\b",
                        r"alianza",
                        r"\bc[aá]mara\b",
                        r"holding",
                        r"grupo",
                        r"constructora",
                        r"\bsimple y por acciones\b",
                        r"\bs\.?c\.?a\.?\b",
                        r"\bs\.?c\.?s\.?\b",
                        r"\bs\.?e\.?\b",
                        r"\bentidad sin animo de lucro\b",
                        r"union temporal",
                        r"inversiones",
                        r"fiduciaria"
                    ]

                    filtrado["es_pj"] = filtrado["Nombre tercero"].astype(str).apply(
                        lambda x: any(re.search(p, norm_txt(x)) for p in patrones_pj)
                    )

                    t1 = tasas[0]["tasa"] if len(tasas) > 0 else None
                    t2 = tasas[1]["tasa"] if len(tasas) > 1 else None

                    ws.cell(row=fila_out, column=1, value=nombre_concepto).font = negro_negrita
                    fila_out += 1
                    fila_inicio = fila_out

                    for _, r in filtrado.iterrows():
                        id_val = str(r["Identificación"]).replace(".0", "")
                        nom_cta = str(r["Nombre cuenta contable"])
                        nom_ter = str(r["Nombre tercero"])
                        cod_cta = str(r["Código cuenta contable"])
                        retencion = float(r["__retencion__"])

                        ws.cell(row=fila_out, column=2, value=cod_cta)
                        ws.cell(row=fila_out, column=3, value=nom_cta)
                        ws.cell(row=fila_out, column=4, value=id_val)
                        ws.cell(row=fila_out, column=5, value=nom_ter)

                        if t1 is not None and t1 != 0:
                            ws.cell(row=fila_out, column=6, value=f"=G{fila_out}/{t1}")
                        else:
                            ws.cell(row=fila_out, column=6, value=0)
                        ws.cell(row=fila_out, column=7, value=retencion if r["es_pj"] else 0)

                        if t2 is not None and t2 != 0:
                            ws.cell(row=fila_out, column=8, value=f"=I{fila_out}/{t2}")
                        else:
                            ws.cell(row=fila_out, column=8, value=0)
                        ws.cell(row=fila_out, column=9, value=retencion if not r["es_pj"] else 0)

                        valores = [ws.cell(row=fila_out, column=c).value for c in range(6, 10)]
                        if all(v in [0, None, ""] for v in valores):
                            continue

                        for col in range(1, 10):
                            c = ws.cell(row=fila_out, column=col)
                            c.border = borde_todo
                            if col >= 6:
                                c.number_format = '#,##0.00'

                        fila_out += 1

                    fila_fin = fila_out - 1
                    ws.cell(row=fila_out, column=1, value=f"Subtotal {nombre_concepto.lower()}").font = negro_negrita
                    for col in range(6, 10):
                        col_letter = get_column_letter(col)
                        formula = f"=SUM({col_letter}{fila_inicio}:{col_letter}{fila_fin})"
                        c = ws.cell(row=fila_out, column=col, value=formula)
                        c.number_format = '#,##0.00'
                        c.font = negro_negrita
                        c.border = borde_todo
                    for col in range(1, 10):
                        c = ws.cell(row=fila_out, column=col)
                        c.fill = gris_subtotal
                        c.font = negro_negrita
                        c.border = borde_todo
                    fila_out += 1
                    return fila_out

                # ==========================================================
                # 🔹 GENERAR TODOS LOS CONCEPTOS
                # ==========================================================
                fila_out = 10

                fila_out = generar_concepto_multitarifa("Rentas de trabajo", ["salario", "pago laboral"], fila_out, tasas=[])
                fila_out = generar_concepto_multitarifa("Rentas de pensiones", ["pension"], fila_out, tasas=[])
                fila_out = generar_concepto_multitarifa("Honorarios", ["honorario", "consultoria"], fila_out, tasas=[{'tasa': 0.11}, {'tasa': 0.10}])
                fila_out = generar_concepto_multitarifa("Comisiones", ["comision", "comisiones"], fila_out, tasas=[{'tasa': 0.11}, {'tasa': 0.10}])
                fila_out = generar_concepto_multitarifa("Servicios 1% (transporte de carga, empresas de servicios temporales (sobre AIU))", ["servicio 1", "servicio 1%", "transporte de carga", "carga", "servicio temporal aiu"], fila_out, tasas=[{'tasa': 0.01}, {'tasa': 0.01}])
                fila_out = generar_concepto_multitarifa("Servicios 2% (vigilancia, aseo, IPS, salud)", ["servicio 2", "vigilancia", "aseo", "vigilancia aiu", "aseo aiu","ips", "salud"], fila_out, tasas=[{'tasa': 0.02}, {'tasa': 0.02}])
                fila_out = generar_concepto_multitarifa("Servicios 3,5% (transporte pasajero, hotel, restaurante, licencia, software)", ["servicio 3,5", "servicio 3.5", "transporte pasajero", "pasajero", "hotel", "restaurante", "licencia", "derecho","software"], fila_out, tasas=[{'tasa': 0.035}, {'tasa': 0.035}])
                fila_out = generar_concepto_multitarifa("Servicios generales", ["servicio 4", "servicio general 4", "servicio declarante", "servicio 6", "servicio general 6", "servicio no declarante", "servicio general no declarante"], fila_out, tasas=[{'tasa': 0.04}, {'tasa': 0.06}])
                fila_out = generar_concepto_multitarifa("Rendimientos financieros e intereses", ["rendimiento", "rendimiento financiero", "intereses"], fila_out, tasas=[{'tasa': 0.07}, {'tasa': 0.07}])
                fila_out = generar_concepto_multitarifa("Arrendamientos muebles", ["arrendamiento mueble", "mueble"], fila_out, tasas=[{'tasa': 0.04}, {'tasa': 0.04}])
                fila_out = generar_concepto_multitarifa("Arrendamientos inmuebles", ["arrendamiento inmueble", "inmueble", "casa", "apartamento"], fila_out, tasas=[{'tasa': 0.035}, {'tasa': 0.035}])
                fila_out = generar_concepto_multitarifa("Regalías y explotación de la propiedad intelectual", ["regalia", "propiedad intelectual", "explitacion"], fila_out, tasas=[{'tasa': 0.02}, {'tasa': 0.02}])
                fila_out = generar_concepto_multitarifa("Dividendos y participaciones", ["dividendo", "participacion"], fila_out, tasas=[])
                fila_out = generar_concepto_multitarifa("Compras generales", ["compra 2,5", "compra 2.5", "compra general 2,5", "compra general 2.5", "compra general declarante", "compra 3,5", "compra 3.5", "compra general 3,5", "compra general 3.5", "compra general no declarante"], fila_out, tasas=[{'tasa': 0.025}, {'tasa': 0.035}])
                fila_out = generar_concepto_multitarifa("Compras combustible", ["combustible", "gasolina"], fila_out, tasas=[{'tasa': 0.001}, {'tasa': 0.001}])
                fila_out = generar_concepto_multitarifa("Compras de bienes raíces para uso vivienda (por las primeras 10.000 UVT)", ["inmueble 1", "bienes raices 1", "menos 10.000 uvt", "vivienda 1"], fila_out, tasas=[{'tasa': 0.01}, {'tasa': 0.01}])
                fila_out = generar_concepto_multitarifa("Compras de bienes raíces para uso vivienda (exceso de las primeras 10.000 UVT)", ["compra inmueble 2,5%", "inmueble 2,5%", "bienes raices 2,5%", "vivienda 2,5%", "compra inmueble 2.5%", "inmueble 2.5%", "bienes raices 2.5%", "vivienda 2.5%", "retencion 2.5%", "mas 10.000 uvt"], fila_out, tasas=[{'tasa': 0.025}, {'tasa': 0.025}])
                fila_out = generar_concepto_multitarifa("Compras de bienes raíces para uso distinto a vivienda de habitación", ["compra inmueble distinto vivienda", "inmueble no habitacion", "bienes raices no habitacion", "vivienda no habitacion", "no habitacion"], fila_out, tasas=[{'tasa': 0.025}, {'tasa': 0.025}])
                fila_out = generar_concepto_multitarifa("Contratos de construcción", ["contrato construccion", "contruccion"], fila_out, tasas=[{'tasa': 0.02}, {'tasa': 0.02}])
                fila_out = generar_concepto_multitarifa("Otros pagos sujetos a retención", ["otros ingresos", "otros pagos", "otros tributacion"], fila_out, tasas=[{'tasa': 0.025}, {'tasa': 0.035}])

                # ==========================================================
                # 🟦 AUTORRETENCIÓN
                # ==========================================================
                def generar_concepto_tarifa_por_nit(concepto, fila_out, datos_cache, nit_empresa):
                    try:
                        if datos_cache and "Tarifas" in datos_cache:
                            df_tarifas = datos_cache["Tarifas"]
                            col_nit = next(c for c in df_tarifas.columns if "nit" in str(c).lower())
                            col_tarifa = next(c for c in df_tarifas.columns if "tarifa" in str(c).lower())
                            fila_tarifa = df_tarifas[df_tarifas[col_nit].astype(str).str.contains(str(nit_empresa), na=False)]
                            tarifa = float(fila_tarifa.iloc[0][col_tarifa]) if not fila_tarifa.empty else None
                        else:
                            tarifa = None
                    except Exception:
                        tarifa = None

                    def normalizar(s):
                        s = str(s).strip().lower()
                        return ''.join(ch for ch in unicodedata.normalize('NFKD', s) if not unicodedata.combining(ch))

                    def buscar_columna(keyword, keyword2=None):
                        for c in raw.columns:
                            nombre = normalizar(c)
                            if keyword in nombre and (keyword2 is None or keyword2 in nombre):
                                return c
                        return None

                    col_nivel = buscar_columna("nivel")
                    col_nombre = buscar_columna("nombre", "cuenta")
                    col_debito = buscar_columna("debito")
                    col_credito = buscar_columna("credito")

                    df_clase_ingresos = raw[
                        raw[col_nivel].astype(str).str.lower().str.contains("clase", na=False)
                        & raw[col_nombre].astype(str).str.lower().str.contains("ingreso", na=False)
                    ]

                    if not df_clase_ingresos.empty:
                        mov_credito = pd.to_numeric(df_clase_ingresos[col_credito], errors="coerce").fillna(0).sum()
                        mov_debito = pd.to_numeric(df_clase_ingresos[col_debito], errors="coerce").fillna(0).sum()
                        base_real = mov_credito - mov_debito
                    else:
                        base_real = 0

                    if tarifa is None or tarifa == 0:
                        base = 0
                        valor_retencion = 0
                    else:
                        base = base_real
                        valor_retencion = f"=F{fila_out}*{tarifa}"

                    ws.cell(row=fila_out, column=1, value=concepto).font = negro_negrita
                    ws.cell(row=fila_out, column=6, value=base)
                    ws.cell(row=fila_out, column=7, value=valor_retencion)
                    ws.cell(row=fila_out, column=8, value=0)
                    ws.cell(row=fila_out, column=9, value=0)

                    for col in range(1, 10):
                        c = ws.cell(row=fila_out, column=col)
                        c.border = borde_todo
                        if col >= 6:
                            c.number_format = '#,##0.00'

                    fila_out += 1

                    ws.cell(row=fila_out, column=1, value="Subtotal autorretención (art. 114-1 E.T.)").font = negro_negrita
                    for col in range(6, 10):
                        letra = get_column_letter(col)
                        formula = f"=SUM({letra}{fila_out-1}:{letra}{fila_out-1})"
                        c = ws.cell(row=fila_out, column=col, value=formula)
                        c.number_format = '#,##0.00'
                        c.font = negro_negrita
                        c.border = borde_todo

                    for col in range(1, 10):
                        c = ws.cell(row=fila_out, column=col)
                        c.fill = gris_subtotal
                        c.font = negro_negrita
                        c.border = borde_todo

                    return fila_out + 1

                fila_out = generar_concepto_tarifa_por_nit(
                    "Autorretención: Contribuyentes exonerados de aportes (art. 114-1 E.T.)",
                    fila_out,
                    datos_generales_cache,
                    nit_numero
                )

                # ==========================================================
                # 🔹 CONCEPTO POST-AUTORRETENCIÓN
                # ==========================================================
                def generar_concepto_post_autorretencion_columnaI(concepto, fila_out):
                    ws.cell(row=fila_out, column=1, value=concepto).font = negro_negrita
                    ws.cell(row=fila_out, column=9, value=0)

                    for col in range(1, 10):
                        c = ws.cell(row=fila_out, column=col)
                        c.border = borde_todo
                        if col == 9:
                            c.number_format = '#,##0.00'

                    fila_out += 1

                    ws.cell(row=fila_out, column=1, value=f"Subtotal {concepto.lower()}").font = negro_negrita
                    c = ws.cell(row=fila_out, column=9, value=f"=SUM(I{fila_out-1}:I{fila_out-1})")
                    c.number_format = '#,##0.00'
                    c.font = negro_negrita
                    c.border = borde_todo

                    for col in range(1, 10):
                        c = ws.cell(row=fila_out, column=col)
                        c.fill = gris_subtotal
                        c.font = negro_negrita
                        c.border = borde_todo

                    return fila_out + 1

                fila_out = generar_concepto_post_autorretencion_columnaI(
                    "Menos retenciones practicadas en exceso o indebidas o por operaciones anuladas, rescindidas o resueltas",
                    fila_out
                )

                # ==========================================================
                # 🔹 TOTAL RETENCIONES RENTA Y COMPLEMENTARIO
                # ==========================================================
                def generar_total_retenciones_renta_y_complementario(fila_out):
                    ws.cell(row=fila_out, column=1, value="Total retenciones renta y complementario").font = negro_negrita

                    filas_subtotales = []
                    fila_exceso = None
                    for row in range(1, fila_out):
                        valor = ws.cell(row=row, column=1).value
                        if valor and isinstance(valor, str):
                            val_norm = valor.lower()
                            if "subtotal" in val_norm:
                                filas_subtotales.append(row)
                            if "exceso" in val_norm or "indebida" in val_norm:
                                fila_exceso = row

                    if filas_subtotales:
                        partes_sumatorias = [f"G{r}" for r in filas_subtotales] + [f"I{r}" for r in filas_subtotales]
                        formula = "=" + "+".join(partes_sumatorias)
                        if fila_exceso:
                            formula += f"-G{fila_exceso}-I{fila_exceso}"
                    else:
                        formula = "=0"

                    cI = ws.cell(row=fila_out, column=9, value=formula)
                    cI.number_format = '#,##0.00'
                    cI.font = negro_negrita

                    for col in range(1, 10):
                        c = ws.cell(row=fila_out, column=col)
                        c.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                        c.font = negro_negrita

                    for col in range(1, 10):
                        c = ws.cell(row=fila_out, column=col)
                        left = Side(style="thin", color="808080") if col == 1 or col == 9 else None
                        right = Side(style="thin", color="808080") if col == 8 or col == 9 else None
                        c.border = Border(
                            left=left,
                            right=right,
                            top=Side(style="thin", color="808080"),
                            bottom=Side(style="thin", color="808080")
                        )

                    return fila_out + 1

                fila_out = generar_total_retenciones_renta_y_complementario(fila_out)

                # ==========================================================
                # 🔹 BLOQUES RETENCIONES OTROS IMPUESTOS
                # ==========================================================
                def generar_bloques_retenciones_otros_impuestos(fila_out):
                    ws.merge_cells(start_row=fila_out, start_column=1, end_row=fila_out, end_column=9)
                    c = ws.cell(row=fila_out, column=1, value="Retenciones practicadas por otros impuestos")
                    c.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                    c.font = negro_negrita
                    c.alignment = Alignment(horizontal="center", vertical="center")

                    for col in range(1, 10):
                        cell = ws.cell(row=fila_out, column=col)
                        cell.border = Border(
                            left=Side(style="thin", color="808080") if col == 1 else None,
                            right=Side(style="thin", color="808080") if col == 9 else None,
                            top=Side(style="thin", color="808080"),
                            bottom=Side(style="thin", color="808080")
                        )

                    fila_out += 1

                    ws.merge_cells(start_row=fila_out, start_column=1, end_row=fila_out, end_column=9)
                    c = ws.cell(row=fila_out, column=1, value="A título de IVA")
                    c.font = negro_negrita
                    c.alignment = Alignment(horizontal="center", vertical="center")

                    for col in range(1, 10):
                        cell = ws.cell(row=fila_out, column=col)
                        cell.border = Border(
                            left=Side(style="thin", color="808080") if col == 1 else None,
                            right=Side(style="thin", color="808080") if col == 9 else None,
                            top=Side(style="thin", color="808080"),
                            bottom=Side(style="thin", color="808080")
                        )

                    return fila_out + 1

                fila_out = generar_bloques_retenciones_otros_impuestos(fila_out)

                # ==========================================================
                # 🔹 RESPONSABLES DEL IMPUESTO SOBRE LAS VENTAS
                # ==========================================================
                def generar_concepto_responsables_impuesto_ventas(fila_out, df_balance):
                    mask_aux = df_balance.apply(es_auxiliar, axis=1)
                    mask_codigo = df_balance["Código cuenta contable"].astype(str).str.startswith("2367")
                    patrones = ["impuesto a las ventas retenido", "retenido 15"]
                    mask_nombre = df_balance["Nombre cuenta contable"].astype(str).apply(lambda x: contiene_alguna(x, patrones))
                    filtrado = df_balance[mask_aux & mask_codigo & mask_nombre].copy()

                    if filtrado.empty:
                        ws.cell(row=fila_out, column=1, value="A responsables del impuesto sobre las ventas").font = negro_negrita
                        ws.cell(row=fila_out, column=9, value=0).number_format = '#,##0.00'
                        for col in range(1, 10):
                            c = ws.cell(row=fila_out, column=col)
                            c.border = Border(
                                left=Side(style="thin", color="808080"),
                                right=Side(style="thin", color="808080"),
                                top=Side(style="thin", color="808080"),
                                bottom=Side(style="thin", color="808080")
                            )
                        return fila_out + 1

                    ws.cell(row=fila_out, column=1, value="A responsables del impuesto sobre las ventas").font = negro_negrita
                    fila_out += 1
                    fila_inicio = fila_out

                    for _, r in filtrado.iterrows():
                        cod_cta = str(r["Código cuenta contable"])
                        nom_cta = str(r["Nombre cuenta contable"])
                        identificacion = str(r.get("Identificación", ""))
                        nom_tercero = str(r.get("Nombre tercero", ""))

                        mov_credito = pd.to_numeric(r.get("Movimiento crédito", 0), errors="coerce") or 0
                        mov_debito = pd.to_numeric(r.get("Movimiento débito", 0), errors="coerce") or 0
                        valor = mov_credito - mov_debito

                        ws.cell(row=fila_out, column=2, value=cod_cta)
                        ws.cell(row=fila_out, column=3, value=nom_cta)
                        ws.cell(row=fila_out, column=4, value=identificacion)
                        ws.cell(row=fila_out, column=5, value=nom_tercero)
                        ws.cell(row=fila_out, column=9, value=valor).number_format = '#,##0.00'

                        for col in range(1, 10):
                            c = ws.cell(row=fila_out, column=col)
                            c.border = Border(
                                left=Side(style="thin", color="808080"),
                                right=Side(style="thin", color="808080"),
                                top=Side(style="thin", color="808080"),
                                bottom=Side(style="thin", color="808080")
                            )

                        fila_out += 1

                    fila_fin = fila_out - 1

                    ws.cell(row=fila_out, column=1, value="Subtotal responsables impuesto sobre las ventas").font = negro_negrita
                    c = ws.cell(row=fila_out, column=9, value=f"=SUM(I{fila_inicio}:I{fila_fin})")
                    c.number_format = '#,##0.00'
                    c.font = negro_negrita
                    for col in range(1, 10):
                        c = ws.cell(row=fila_out, column=col)
                        c.fill = gris_subtotal
                        c.border = Border(
                            left=Side(style="thin", color="808080"),
                            right=Side(style="thin", color="808080"),
                            top=Side(style="thin", color="808080"),
                            bottom=Side(style="thin", color="808080")
                        )

                    return fila_out + 1

                fila_out = generar_concepto_responsables_impuesto_ventas(fila_out, raw)

                # ==========================================================
                # 🔹 NO RESIDENTES
                # ==========================================================
                def generar_concepto_no_residentes(fila_out):
                    ws.cell(row=fila_out, column=1, value="Practicadas por servicios a no residentes o no domiciliados").font = negro_negrita
                    ws.cell(row=fila_out, column=9, value=0).number_format = '#,##0.00'

                    for col in range(1, 10):
                        c = ws.cell(row=fila_out, column=col)
                        c.border = Border(
                            left=Side(style="thin", color="808080"),
                            right=Side(style="thin", color="808080"),
                            top=Side(style="thin", color="808080"),
                            bottom=Side(style="thin", color="808080")
                        )

                    fila_out += 1

                    ws.cell(row=fila_out, column=1, value="Subtotal servicios a no residentes o no domiciliados").font = negro_negrita
                    c = ws.cell(row=fila_out, column=9, value=f"=SUM(I{fila_out-1}:I{fila_out-1})")
                    c.number_format = '#,##0.00'
                    c.font = negro_negrita

                    for col in range(1, 10):
                        c = ws.cell(row=fila_out, column=col)
                        c.fill = gris_subtotal
                        c.border = Border(
                            left=Side(style="thin", color="808080"),
                            right=Side(style="thin", color="808080"),
                            top=Side(style="thin", color="808080"),
                            bottom=Side(style="thin", color="808080")
                        )

                    return fila_out + 1

                fila_out = generar_concepto_no_residentes(fila_out)

                # ==========================================================
                # 🔹 RETENCIONES EN EXCESO (IVA)
                # ==========================================================
                def generar_concepto_retenciones_exceso(fila_out):
                    ws.cell(row=fila_out, column=1, value="Menos retenciones practicadas en exceso o indebidas o por operaciones anuladas, rescindidas o resueltas").font = negro_negrita
                    ws.cell(row=fila_out, column=9, value=0).number_format = '#,##0.00'

                    for col in range(1, 10):
                        c = ws.cell(row=fila_out, column=col)
                        c.border = Border(
                            left=Side(style="thin", color="808080"),
                            right=Side(style="thin", color="808080"),
                            top=Side(style="thin", color="808080"),
                            bottom=Side(style="thin", color="808080")
                        )

                    fila_out += 1

                    ws.cell(row=fila_out, column=1, value="Subtotal retenciones practicadas en exceso o indebidas").font = negro_negrita
                    c = ws.cell(row=fila_out, column=9, value=f"=SUM(I{fila_out-1}:I{fila_out-1})")
                    c.number_format = '#,##0.00'
                    c.font = negro_negrita

                    for col in range(1, 10):
                        c = ws.cell(row=fila_out, column=col)
                        c.fill = gris_subtotal
                        c.border = Border(
                            left=Side(style="thin", color="808080"),
                            right=Side(style="thin", color="808080"),
                            top=Side(style="thin", color="808080"),
                            bottom=Side(style="thin", color="808080")
                        )

                    return fila_out + 1

                fila_out = generar_concepto_retenciones_exceso(fila_out)

                # ==========================================================
                # 🔹 TOTAL RETENCIONES IVA
                # ==========================================================
                def generar_total_retenciones_iva(fila_out):
                    ws.cell(row=fila_out, column=1, value="Total retenciones IVA").font = negro_negrita

                    fila_responsables, fila_no_residentes, fila_exceso = None, None, None
                    for row in range(1, fila_out):
                        valor = ws.cell(row=row, column=1).value
                        if not valor or not isinstance(valor, str):
                            continue
                        val_norm = valor.lower()
                        if "subtotal responsables impuesto sobre las ventas" in val_norm:
                            fila_responsables = row
                        elif "subtotal servicios a no residentes" in val_norm:
                            fila_no_residentes = row
                        elif "subtotal retenciones practicadas en exceso" in val_norm or "indebidas" in val_norm:
                            fila_exceso = row

                    partes = []
                    if fila_responsables:
                        partes.append(f"I{fila_responsables}")
                    if fila_no_residentes:
                        partes.append(f"I{fila_no_residentes}")
                    formula = "=" + "+".join(partes)
                    if fila_exceso:
                        formula += f"-I{fila_exceso}"
                    if not partes:
                        formula = "=0"

                    cI = ws.cell(row=fila_out, column=9, value=formula)
                    cI.number_format = '#,##0.00'
                    cI.font = negro_negrita

                    for col in range(1, 10):
                        c = ws.cell(row=fila_out, column=col)
                        c.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                        c.font = negro_negrita

                    for col in range(1, 10):
                        c = ws.cell(row=fila_out, column=col)
                        if 1 <= col <= 8:
                            left = Side(style="thin", color="808080") if col == 1 else None
                            right = Side(style="thin", color="808080") if col == 8 else None
                            c.border = Border(
                                left=left,
                                right=right,
                                top=Side(style="thin", color="808080"),
                                bottom=Side(style="thin", color="808080")
                            )
                        elif col == 9:
                            c.border = Border(
                                left=Side(style="thin", color="808080"),
                                right=Side(style="thin", color="808080"),
                                top=Side(style="thin", color="808080"),
                                bottom=Side(style="thin", color="808080")
                            )

                    return fila_out + 1

                fila_out = generar_total_retenciones_iva(fila_out)

                # ==========================================================
                # 🔹 TOTAL RETENCIONES
                # ==========================================================
                def generar_total_retenciones(fila_out):
                    ws.cell(row=fila_out, column=1, value="Total retenciones").font = negro_negrita

                    fila_total_renta, fila_total_iva = None, None
                    for row in range(1, fila_out):
                        valor = ws.cell(row=row, column=1).value
                        if not valor or not isinstance(valor, str):
                            continue
                        val_norm = valor.lower()
                        if "total retenciones renta y complementario" in val_norm:
                            fila_total_renta = row
                        elif "total retenciones iva" in val_norm:
                            fila_total_iva = row

                    partes = []
                    if fila_total_renta:
                        partes.append(f"I{fila_total_renta}")
                    if fila_total_iva:
                        partes.append(f"I{fila_total_iva}")
                    formula = "=" + "+".join(partes) if partes else "=0"

                    cI = ws.cell(row=fila_out, column=9, value=formula)
                    cI.number_format = '#,##0.00'
                    cI.font = negro_negrita

                    for col in range(1, 10):
                        c = ws.cell(row=fila_out, column=col)
                        c.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                        c.font = negro_negrita

                    for col in range(1, 10):
                        c = ws.cell(row=fila_out, column=col)
                        if 1 <= col <= 8:
                            left = Side(style="thin", color="808080") if col == 1 else None
                            right = Side(style="thin", color="808080") if col == 8 else None
                            c.border = Border(
                                left=left,
                                right=right,
                                top=Side(style="thin", color="808080"),
                                bottom=Side(style="thin", color="808080")
                            )
                        elif col == 9:
                            c.border = Border(
                                left=Side(style="thin", color="808080"),
                                right=Side(style="thin", color="808080"),
                                top=Side(style="thin", color="808080"),
                                bottom=Side(style="thin", color="808080")
                            )

                    return fila_out + 1

                fila_out = generar_total_retenciones(fila_out)

                # ==========================================================
                # 🔹 SANCIONES
                # ==========================================================
                def generar_concepto_sanciones(fila_out):
                    ws.cell(row=fila_out, column=1, value="Sanciones").font = negro_negrita

                    cI = ws.cell(row=fila_out, column=9, value=0)
                    cI.number_format = '#,##0.00'
                    cI.font = negro_negrita

                    for col in range(1, 10):
                        c = ws.cell(row=fila_out, column=col)
                        c.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                        c.font = negro_negrita

                    for col in range(1, 10):
                        c = ws.cell(row=fila_out, column=col)
                        if 1 <= col <= 8:
                            left = Side(style="thin", color="808080") if col == 1 else None
                            right = Side(style="thin", color="808080") if col == 8 else None
                            c.border = Border(
                                left=left,
                                right=right,
                                top=Side(style="thin", color="808080"),
                                bottom=Side(style="thin", color="808080")
                            )
                        elif col == 9:
                            c.border = Border(
                                left=Side(style="thin", color="808080"),
                                right=Side(style="thin", color="808080"),
                                top=Side(style="thin", color="808080"),
                                bottom=Side(style="thin", color="808080")
                            )

                    return fila_out + 1

                fila_out = generar_concepto_sanciones(fila_out)

                # ==========================================================
                # 🔹 TOTAL RETENCIONES MÁS SANCIONES
                # ==========================================================
                def generar_total_retenciones_mas_sanciones(fila_out):
                    ws.cell(row=fila_out, column=1, value="Total retenciones más sanciones").font = negro_negrita

                    fila_total_ret, fila_sanciones = None, None
                    for row in range(1, fila_out):
                        valor = ws.cell(row=row, column=1).value
                        if not valor or not isinstance(valor, str):
                            continue
                        val_norm = valor.lower()
                        if "total retenciones" in val_norm and "más sanciones" not in val_norm:
                            fila_total_ret = row
                        elif "sanciones" in val_norm and "total" not in val_norm:
                            fila_sanciones = row

                    partes = []
                    if fila_total_ret:
                        partes.append(f"I{fila_total_ret}")
                    if fila_sanciones:
                        partes.append(f"I{fila_sanciones}")
                    formula = "=" + "+".join(partes) if partes else "=0"

                    cI = ws.cell(row=fila_out, column=9, value=formula)
                    cI.number_format = '#,##0.00'
                    cI.font = negro_negrita

                    for col in range(1, 10):
                        c = ws.cell(row=fila_out, column=col)
                        c.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                        c.font = negro_negrita

                    for col in range(1, 10):
                        c = ws.cell(row=fila_out, column=col)
                        if 1 <= col <= 8:
                            left = Side(style="thin", color="808080") if col == 1 else None
                            right = Side(style="thin", color="808080") if col == 8 else None
                            c.border = Border(
                                left=left,
                                right=right,
                                top=Side(style="thin", color="808080"),
                                bottom=Side(style="thin", color="808080")
                            )
                        elif col == 9:
                            c.border = Border(
                                left=Side(style="thin", color="808080"),
                                right=Side(style="thin", color="808080"),
                                top=Side(style="thin", color="808080"),
                                bottom=Side(style="thin", color="808080")
                            )

                    return fila_out + 1

                fila_out = generar_total_retenciones_mas_sanciones(fila_out)

                # ==========================================================
                # 👻 LIMPIEZA: OCULTAR FILAS CON CEROS
                # ==========================================================
                def ocultar_filas_con_formulas_cero(ws):
                    fila_inicio, fila_fin = None, None

                    for row in range(1, ws.max_row + 1):
                        val = ws.cell(row=row, column=1).value
                        if val and isinstance(val, str):
                            texto = val.lower()
                            if "rentas de trabajo" in texto and fila_inicio is None:
                                fila_inicio = row
                            elif "autorretencion" in texto or "autorretención" in texto or "autorreteción" in texto:
                                fila_fin = row
                                break

                    if not fila_inicio or not fila_fin:
                        return

                    def es_cero_o_formula_cero(v):
                        if v is None:
                            return False
                        v_str = str(v).strip().lower().replace(",", ".")
                        if v_str in ["0", "0.0", "0.00", "=0"]:
                            return True
                        if v_str.startswith("=") and any(pat in v_str for pat in ["/0.", "*0", "=0"]):
                            return True
                        return False

                    for row in range(fila_inicio + 1, fila_fin):
                        datos_BE = [ws.cell(row=row, column=col).value for col in range(2, 6)]
                        valores_FI = [ws.cell(row=row, column=col).value for col in range(6, 10)]

                        hay_datos_BE = any(datos_BE)
                        todos_ceros_FI = all(es_cero_o_formula_cero(v) for v in valores_FI)

                        if hay_datos_BE and todos_ceros_FI:
                            ws.row_dimensions[row].hidden = True

                ocultar_filas_con_formulas_cero(ws)

                # ==========================================================
                # 🧩 COPIAR HOJA Y CREAR FORMULARIO_350
                # ==========================================================
                def generar_formulario_350_inicial(wb):
                    if "Detalle_350" not in wb.sheetnames:
                        return

                    ws_origen = wb["Detalle_350"]
                    ws_nueva = wb.create_sheet("Formulario_350")

                    encabezados = [
                        "Fila técnica", "Columna B", "Columna C", "Columna D",
                        "Columna E", "Columna F", "Columna G", "Columna H", "Columna I"
                    ]
                    for col, texto in enumerate(encabezados, start=1):
                        ws_nueva.cell(row=1, column=col, value=texto).font = Font(bold=True, color="000000")

                    for row in ws_origen.iter_rows():
                        for cell in row:
                            try:
                                new_cell = ws_nueva.cell(row=cell.row + 1, column=cell.col_idx, value=cell.value)
                                if cell.has_style:
                                    new_cell.font = copy(cell.font)
                                    new_cell.fill = copy(cell.fill)
                                    new_cell.border = copy(cell.border)
                                    new_cell.alignment = copy(cell.alignment)
                                    new_cell.number_format = cell.number_format
                            except AttributeError:
                                continue

                    for merged_range in ws_origen.merged_cells.ranges:
                        coords = str(merged_range)
                        letras, numeros = [], []
                        for c in coords.replace(":", " ").split():
                            col = ''.join([ch for ch in c if ch.isalpha()])
                            row = ''.join([ch for ch in c if ch.isdigit()])
                            letras.append(col)
                            numeros.append(int(row) + 1)
                        rango_nuevo = f"{letras[0]}{numeros[0]}:{letras[1]}{numeros[1]}"
                        try:
                            ws_nueva.merge_cells(rango_nuevo)
                        except ValueError:
                            pass

                    for col_letter, dim in ws_origen.column_dimensions.items():
                        ws_nueva.column_dimensions[col_letter].width = dim.width if dim.width else 12

                    for row_idx, dim in ws_origen.row_dimensions.items():
                        if dim.height:
                            ws_nueva.row_dimensions[row_idx + 1].height = dim.height

                    for row_dim in ws_nueva.row_dimensions.values():
                        row_dim.hidden = False

                    ws_nueva.column_dimensions.group("B", "E", hidden=True, outline_level=1)
                    ws_nueva.sheet_view.showOutlineSymbols = True
                    ws_nueva.sheet_view.outline_summary_right = True
                    ws_nueva.sheet_view.showRowColHeaders = True

                generar_formulario_350_inicial(wb)

                # ==========================================================
                # 🧩 AGRUPAR Y COLAPSAR FILAS
                # ==========================================================
                def agrupar_colapsar_formulario_350(wb):
                    if "Formulario_350" not in wb.sheetnames:
                        return

                    ws = wb["Formulario_350"]

                    def encontrar_filas(ws, inicio_texto, fin_texto):
                        fila_inicio, fila_fin = None, None
                        for row in range(1, ws.max_row + 1):
                            val = ws.cell(row=row, column=1).value
                            if not val or not isinstance(val, str):
                                continue
                            texto = val.lower()
                            if inicio_texto in texto and fila_inicio is None:
                                fila_inicio = row
                            if fin_texto in texto:
                                fila_fin = row
                                break
                        return fila_inicio, fila_fin

                    def aplicar_colapso(ws, fila_inicio, fila_fin):
                        if not fila_inicio or not fila_fin:
                            return

                        for row in range(fila_inicio, fila_fin + 1):
                            val = ws.cell(row=row, column=1).value
                            texto = str(val).strip().lower() if val else ""

                            if texto == "" or not texto.startswith("subtotal"):
                                ws.row_dimensions[row].outlineLevel = 1
                                ws.row_dimensions[row].hidden = True

                    inicio1, fin1 = encontrar_filas(
                        ws,
                        "rentas de trabajo",
                        "subtotal menos retenciones practicadas en exceso"
                    )
                    aplicar_colapso(ws, inicio1, fin1)

                    inicio2, fin2 = encontrar_filas(
                        ws,
                        "a responsables del impuesto sobre las ventas",
                        "subtotal retenciones practicadas en exceso"
                    )
                    aplicar_colapso(ws, inicio2, fin2)

                    ws.sheet_view.showOutlineSymbols = True
                    ws.sheet_view.outline_summary_below = True

                agrupar_colapsar_formulario_350(wb)

                # ==========================================================
                # 🧩 AJUSTE DE FÓRMULAS (+1 FILA)
                # ==========================================================
                def ajustar_todas_las_referencias(wb, hoja="Formulario_350"):
                    if hoja not in wb.sheetnames:
                        return

                    ws = wb[hoja]
                    patron = re.compile(r"(\$?[A-Z]+)(\d+)")

                    for row in ws.iter_rows():
                        for cell in row:
                            valor = cell.value
                            if isinstance(valor, str) and valor.strip().startswith("="):
                                nuevo_valor = patron.sub(
                                    lambda m: f"{m.group(1)}{int(m.group(2)) + 1}", valor
                                )
                                if nuevo_valor != valor:
                                    cell.value = nuevo_valor

                ajustar_todas_las_referencias(wb)

                # ==========================================================
                # 🧩 REESCRIBIR FÓRMULAS CON PARÉNTESIS
                # ==========================================================
                def reescribir_formulas_con_parentesis_form350(wb):
                    hoja = "Formulario_350"

                    if hoja not in wb.sheetnames:
                        return

                    ws = wb[hoja]

                    for row in ws.iter_rows():
                        for cell in row:
                            valor = cell.value
                            if isinstance(valor, str) and valor.strip().startswith("="):
                                contenido = valor[1:].strip()
                                if not (contenido.startswith("(") and contenido.endswith(")")):
                                    cell.value = f"=({contenido})"

                reescribir_formulas_con_parentesis_form350(wb)

                # ==========================================================
                # 🧩 REDONDEAR AL MÚLTIPLO DE MIL
                # ==========================================================
                def redondear_multiplo_mil_formulario_350(wb):
                    hoja = "Formulario_350"
                    if hoja not in wb.sheetnames:
                        return

                    ws = wb[hoja]

                    for row in ws.iter_rows():
                        for cell in row:
                            valor = cell.value
                            if isinstance(valor, str) and valor.strip().startswith("="):
                                contenido = valor[1:].strip()
                                if "MROUND" not in contenido.upper():
                                    cell.value = f"=MROUND({contenido},1000)"
                            elif isinstance(valor, (int, float)) and valor != 0:
                                cell.value = f"=MROUND({valor},1000)"

                redondear_multiplo_mil_formulario_350(wb)

                # ==========================================================
                # 🧩 AGREGAR BALANCE ORIGINAL (desde fila 8 en adelante)
                # ==========================================================
                def agregar_hoja_balance_original(wb, balance_file):
                    try:
                        balance_file.seek(0)
                        wb_balance = openpyxl.load_workbook(balance_file, data_only=False)
                        hoja_origen = wb_balance.active
                        hoja_destino = wb.create_sheet("Anexo_Balance_Original")

                        FILA_INICIO = 8  # Empezar desde la fila 8

                        # 1️⃣ Identificar celdas combinadas que están en la fila 8 o posterior
                        merged_ranges_to_copy = []
                        for merged_range in hoja_origen.merged_cells.ranges:
                            # Solo copiar si la celda combinada comienza en fila 8 o después
                            if merged_range.min_row >= FILA_INICIO:
                                merged_ranges_to_copy.append(merged_range)

                        # 2️⃣ Copiar valores y estilos (desde fila 8)
                        for row in hoja_origen.iter_rows(min_row=FILA_INICIO):
                            for cell in row:
                                # Saltar celdas combinadas (MergedCell)
                                if isinstance(cell, openpyxl.cell.cell.MergedCell):
                                    continue

                                try:
                                    # Calcular nueva posición (restar 7 para que fila 8 sea fila 1)
                                    new_row = cell.row - (FILA_INICIO - 1)
                                    new_col = cell.column
                                    new_cell = hoja_destino.cell(row=new_row, column=new_col)

                                    # Copiar valor
                                    if isinstance(cell.value, (int, float, str)):
                                        new_cell.value = cell.value
                                    elif cell.value is not None:
                                        new_cell.value = str(cell.value)

                                    # Copiar estilos si existen
                                    if cell.has_style:
                                        try:
                                            new_cell.font = copy(cell.font)
                                            new_cell.fill = copy(cell.fill)
                                            new_cell.border = copy(cell.border)
                                            new_cell.alignment = copy(cell.alignment)
                                            new_cell.number_format = cell.number_format
                                        except:
                                            pass
                                except Exception:
                                    continue

                        # 3️⃣ Copiar dimensiones de columnas
                        for col_letter, dim in hoja_origen.column_dimensions.items():
                            try:
                                hoja_destino.column_dimensions[col_letter].width = dim.width if dim.width else 12
                            except:
                                pass

                        # 4️⃣ Copiar dimensiones de filas (solo desde fila 8)
                        for row_idx, dim in hoja_origen.row_dimensions.items():
                            if row_idx >= FILA_INICIO and dim.height:
                                try:
                                    new_row_idx = row_idx - (FILA_INICIO - 1)
                                    hoja_destino.row_dimensions[new_row_idx].height = dim.height
                                except:
                                    pass

                        # 5️⃣ Aplicar celdas combinadas (ajustadas a la nueva posición)
                        for merged_range in merged_ranges_to_copy:
                            try:
                                # Ajustar posiciones: restar 7 para que fila 8 sea fila 1
                                start_row = merged_range.min_row - (FILA_INICIO - 1)
                                end_row = merged_range.max_row - (FILA_INICIO - 1)
                                start_col = merged_range.min_col
                                end_col = merged_range.max_col

                                # Solo aplicar si la fila ajustada es válida (> 0)
                                if start_row > 0 and end_row > 0:
                                    hoja_destino.merge_cells(
                                        start_row=start_row,
                                        start_column=start_col,
                                        end_row=end_row,
                                        end_column=end_col
                                    )
                            except Exception:
                                continue

                        # ✅ Mensaje de éxito
                        st.success(f"✅ Hoja 'Anexo_Balance_Original' agregada correctamente desde fila {FILA_INICIO}")

                    except Exception as e:
                        st.warning(f"⚠️ No se pudo copiar el balance original: {e}")

                agregar_hoja_balance_original(wb, balance_file)

                # ==========================================================
                # 🧩 AJUSTES FINALES EN FORMULARIO_350
                # ==========================================================
                def ajustar_formulario_350_final(wb, nombre_empresa):
                    if "Formulario_350" not in wb.sheetnames:
                        return

                    ws = wb["Formulario_350"]

                    for row in range(1, ws.max_row + 1):
                        valor = ws.cell(row=row, column=1).value
                        if not valor or not isinstance(valor, str):
                            continue

                        texto = valor.strip()
                        if texto.lower().startswith("subtotal "):
                            nuevo_texto = texto[len("Subtotal "):].strip()
                            if nuevo_texto:
                                nuevo_texto = nuevo_texto[0].upper() + nuevo_texto[1:]
                            ws.cell(row=row, column=1, value=nuevo_texto)

                            for col in range(1, 10):
                                celda = ws.cell(row=row, column=col)
                                celda.fill = PatternFill(fill_type=None)
                                celda.font = Font(
                                    name=celda.font.name,
                                    size=celda.font.size,
                                    color=celda.font.color,
                                    bold=False,
                                    italic=celda.font.italic
                                )

                    ws.row_dimensions[1].hidden = True
                    wb._sheets.sort(key=lambda s: 0 if s.title == "Formulario_350" else 1)

                ajustar_formulario_350_final(wb, nombre_empresa)

                # ==========================================================
                # 💾 GUARDAR ARCHIVO EN MEMORIA
                # ==========================================================
                nombre_limpio = "".join(ch for ch in nombre_empresa if ch.isalnum() or ch in (" ", "_", "-")).strip()
                if not nombre_limpio:
                    nombre_limpio = f"Empresa_{idx}"

                out_name = f"Formulario_350_{nombre_limpio}.xlsx"

                # Guardar en BytesIO
                output = BytesIO()
                wb.save(output)
                output.seek(0)

                archivos_generados.append({
                    'nombre': out_name,
                    'datos': output.getvalue(),
                    'empresa': nombre_empresa
                })

                archivos_procesados += 1
                progress_bar.progress(archivos_procesados / archivos_totales)

            except Exception as e:
                st.error(f"❌ ERROR procesando {balance_file.name}: {str(e)}")
                st.code(traceback.format_exc())
                continue

        # ==========================================================
        # 🎉 MOSTRAR RESULTADOS Y DESCARGAS
        # ==========================================================
        status_text.empty()
        progress_bar.empty()

        st.markdown("---")
        st.success(f"🎉 Procesamiento completado: {archivos_procesados}/{archivos_totales} archivos")

        if archivos_generados:
            st.subheader("📥 Descargar Formularios Generados")

            cols = st.columns(min(3, len(archivos_generados)))

            for i, archivo in enumerate(archivos_generados):
                with cols[i % 3]:
                    st.download_button(
                        label=f"📄 {archivo['empresa']}",
                        data=archivo['datos'],
                        file_name=archivo['nombre'],
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )

        if archivos_procesados < archivos_totales:
            st.warning(f"⚠️ {archivos_totales - archivos_procesados} archivo(s) con errores")

        # ==========================================================
        # 📦 DESCARGA DE TODOS LOS FORMULARIOS EN UN ZIP
        # ==========================================================
        if archivos_generados:
            st.markdown("---")
        
            # Crear ZIP en memoria
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zipf:
                for archivo in archivos_generados:
                    zipf.writestr(archivo["nombre"], archivo["datos"])
            zip_buffer.seek(0)
        
            # Botón principal
            st.download_button(
                label=f"📦 Descargar todos los formularios ({len(archivos_generados)} archivos)",
                data=zip_buffer,
                file_name="Formularios_350_Completos.zip",
                mime="application/zip",
                use_container_width=True,
                type="primary",
            )
        
            # Lista desplegable
            with st.expander("📋 Ver archivos generados"):
                for i, archivo in enumerate(archivos_generados, start=1):
                    st.markdown(f"**{i}. {archivo['empresa']}**")

# ==========================================================
# 📖 INFORMACIÓN ADICIONAL
# ==========================================================
with st.expander("📘 Instrucciones de Uso"):
    st.markdown("""
    ### 📋 Cómo usar este generador:

    1. **Archivos de Balance**: Sube uno o varios archivos Excel con los balances contables de tus empresas
       - Deben tener el formato estándar con encabezado en la fila 8
       - Deben contener las columnas: Nivel, Transaccional, Código cuenta contable, Nombre cuenta contable, etc.
       
    <br>
    
    2. **Datos Generales**: Sube un archivo Excel con dos hojas:
       - **Hoja "Tarifas"**: Debe contener las columnas NIT, CIIU, Actividad económica y Tarifa
       - Los NITs deben coincidir con los de los archivos de balance
       
    <br>

    3. **Procesar**: Haz clic en el botón "Generar Formularios 350"
           
    <br>
    
    4. **Descargar**: Una vez procesado, descarga cada formulario generado

    ### 📊 El archivo generado contiene:

    - **Formulario_350**: Hoja principal con formato colapsado y fórmulas ajustadas
    - **Detalle_350**: Hoja con el detalle completo de todos los conceptos
    - **Anexo_Balance_Original**: Copia del balance original usado para el cálculo

    ### ⚠️ Requisitos:

    - Los archivos de balance deben seguir el formato estándar de Siigo/WorldOffice
    - La información de empresa, NIT y período debe estar en las primeras filas
    - Las cuentas contables deben seguir el PUC colombiano
    """, unsafe_allow_html=True)

with st.expander("🧩 Características del Formulario"):
    st.markdown("""
    ### ✨ Funcionalidades incluidas:

    - Cálculo automático de retenciones por concepto
    - Clasificación automática entre personas jurídicas y naturales
    - Búsqueda de tarifas personalizadas por NIT
    - Cálculo de autorretención (Art. 114-1 E.T.)
    - Inclusión de retenciones de IVA
    - Filas ocultas automáticamente cuando no hay movimiento
    - Formato profesional con colores y bordes
    - Redondeo automático al múltiplo de mil
    - Fórmulas optimizadas para Excel

    ### 📋 Conceptos procesados:

    - Rentas de trabajo y pensiones
    - Honorarios y comisiones
    - Servicios (1%, 2%, 3.5%, generales)
    - Rendimientos financieros
    - Arrendamientos (muebles e inmuebles)
    - Compras generales y especiales
    - Y muchos más...
    """)
