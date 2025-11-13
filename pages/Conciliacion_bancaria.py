import streamlit as st
import pandas as pd
import unicodedata
import datetime
import re
import io
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from ui_utils import aplicar_css_global

# ==========================================================
# 🔐 VERIFICACIÓN DE AUTENTICACIÓN
# ==========================================================
if 'autenticado' not in st.session_state:
    st.session_state.autenticado = False

if not st.session_state.autenticado:
    st.error("🔒 Debes iniciar sesión primero")
    st.info("👈 Ve a la página principal para autenticarte")
    st.stop()

# ==========================================================
# 🎨 APLICAR ESTILOS GLOBALES
# ==========================================================
aplicar_css_global()

# ==========================================================
# 🏛 CONFIGURACIÓN DE PÁGINA
# ==========================================================
st.set_page_config(
    page_title="Conciliación Bancaria",
    layout="wide",
    page_icon="🏦"
)

st.title("🏦 Generador de Conciliación Bancaria")
st.markdown("---")

# ==========================================================
# 🧭 SIDEBAR CON INFORMACIÓN DE USUARIO
# ==========================================================
with st.sidebar:
    st.markdown("---")
    st.success(f"👤 Usuario: **{st.session_state.username}**")
    if st.button("🚪 Cerrar Sesión", key="logout_conciliacion"):
        st.session_state.autenticado = False
        st.session_state.username = None
        st.switch_page("Home.py")

# ==========================================================
# 🔧 FUNCIONES AUXILIARES
# ==========================================================
def limpiar_columna(col):
    col = col.strip().lower()
    col = "".join(c for c in unicodedata.normalize("NFD", col) if unicodedata.category(c) != "Mn")
    return col.replace("  ", " ")

def buscar_columna(df, palabra):
    cols = [c for c in df.columns if palabra in c]
    if not cols:
        return None
    return cols[0]

def cargar_archivo(uploaded_file, header=None):
    """Carga archivo desde Streamlit uploader"""
    if uploaded_file.name.endswith(".csv"):
        return pd.read_csv(uploaded_file, header=header)
    else:
        return pd.read_excel(uploaded_file, header=header)

def convertir_fecha_unificada(col):
    def _convert(val):
        if pd.isna(val):
            return None

        if isinstance(val, (datetime.datetime, pd.Timestamp)):
            return val

        if isinstance(val, (int, float)):
            if val > 1000:
                try:
                    return datetime.datetime(1899, 12, 30) + datetime.timedelta(days=val)
                except:
                    pass
            if 1900 <= val <= 2100:
                try:
                    return datetime.datetime(int(val), 1, 1)
                except:
                    pass

        if isinstance(val, str):
            val = val.strip()
            formatos = [
                "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y", "%Y/%m/%d",
                "%d.%m.%Y", "%d %m %Y", "%Y%m%d", "%d/%m/%y", "%d-%m-%y",
                "%d/%m/%Y %H:%M:%S", "%Y-%m-%d %H:%M:%S"
            ]
            for fmt in formatos:
                try:
                    return datetime.datetime.strptime(val, fmt)
                except:
                    continue
            try:
                return pd.to_datetime(val, dayfirst=True, errors="coerce")
            except:
                pass
        return None
    return col.apply(_convert)

def identificar_ing_gas(desc, prefijos):
    if not isinstance(desc, str):
        return None
    for pref in prefijos:
        if desc.startswith(pref.lower()):
            return pref
    return None

# ==========================================================
# 📤 INTERFAZ DE CARGA DE ARCHIVOS
# ==========================================================
col1, col2 = st.columns(2)

with col1:
    st.subheader("📊 Archivo de Contabilidad")
    uploaded_contab = st.file_uploader(
        "Sube el archivo 'Movimiento auxiliar por cuenta contable'",
        type=["xlsx", "csv"],
        key="contabilidad",
        help="Archivo Excel o CSV con el detalle contable del banco"
    )

with col2:
    st.subheader("🏦 Extracto Bancario")
    uploaded_banco = st.file_uploader(
        "Sube el extracto bancario",
        type=["xlsx", "csv"],
        key="banco",
        help="Archivo Excel o CSV con los movimientos del banco"
    )

st.markdown("---")

# ==========================================================
# ⚙️ BOTÓN DE PROCESAMIENTO
# ==========================================================
if st.button("⚙️ Generar Conciliación Bancaria", type="primary", use_container_width=True):
    
    # Validar archivos cargados
    if not uploaded_contab:
        st.error("⚠️ Por favor sube el archivo de contabilidad")
        st.stop()
    
    if not uploaded_banco:
        st.error("⚠️ Por favor sube el extracto bancario")
        st.stop()

    with st.spinner("⏳ Procesando archivos..."):
        try:
            # ==========================================================
            # 📊 CARGA Y NORMALIZACIÓN DE ARCHIVOS
            # ==========================================================
            
            # Cargar contabilidad
            df_contable_raw = cargar_archivo(uploaded_contab, header=7)
            df_contable_raw.columns = [limpiar_columna(c) for c in df_contable_raw.columns]
            st.success(f"✅ Archivo de contabilidad cargado: {len(df_contable_raw)} registros")
            
            # Cargar banco
            df_banco_raw = cargar_archivo(uploaded_banco, header=0)
            df_banco_raw.columns = [limpiar_columna(c) for c in df_banco_raw.columns]
            st.success(f"✅ Extracto bancario cargado: {len(df_banco_raw)} registros")

            # ==========================================================
            # 🔄 TRANSFORMACIÓN CONTABILIDAD
            # ==========================================================
            
            # Detectar columnas
            col_fecha = None
            for c in df_contable_raw.columns:
                if "fecha" in c.lower():
                    col_fecha = c
                    break
            
            col_comp = buscar_columna(df_contable_raw, "comprobante") or df_contable_raw.columns[0]
            col_tercero = buscar_columna(df_contable_raw, "tercero") or df_contable_raw.columns[1]
            col_debito = buscar_columna(df_contable_raw, "debito") or df_contable_raw.columns[-2]
            col_credito = buscar_columna(df_contable_raw, "credito") or df_contable_raw.columns[-1]
            
            # Convertir fechas
            if col_fecha:
                df_contable_raw["__fecha_tmp__"] = convertir_fecha_unificada(df_contable_raw[col_fecha])
                st.info(f"📅 Columna de fecha detectada: '{col_fecha}'")
            else:
                st.warning("⚠️ No se encontró columna de fecha en contabilidad")
                df_contable_raw["__fecha_tmp__"] = None
            
            # Filtrar por última fecha válida
            df_contable_filtrado = df_contable_raw[df_contable_raw["__fecha_tmp__"].notna()].copy()
            
            if not df_contable_filtrado.empty:
                ultima_fecha_valida = df_contable_filtrado["__fecha_tmp__"].max()
                indice_ultima_fecha = df_contable_filtrado[df_contable_filtrado["__fecha_tmp__"] == ultima_fecha_valida].index.max()
                df_contable_filtrado = df_contable_filtrado.loc[:indice_ultima_fecha].copy()
                st.info(f"📆 Última fecha contable: {ultima_fecha_valida.strftime('%d/%m/%Y')}")
            
            # Crear DataFrame contable final
            df_contable = pd.DataFrame()
            df_contable["fecha"] = df_contable_filtrado["__fecha_tmp__"]
            df_contable["descripcion"] = (
                df_contable_filtrado[col_comp].astype(str).str.strip() + " " +
                df_contable_filtrado[col_tercero].astype(str).str.strip()
            ).str.lower()
            df_contable["monto"] = (
                pd.to_numeric(df_contable_filtrado[col_debito], errors="coerce").fillna(0) -
                pd.to_numeric(df_contable_filtrado[col_credito], errors="coerce").fillna(0)
            )
            df_contable = df_contable.dropna(subset=["monto"])

            # ==========================================================
            # 🏦 TRANSFORMACIÓN BANCOS
            # ==========================================================
            
            dicc_columnas = {
                "fecha": ["fecha", "fecha operacion", "dia"],
                "descripcion": ["descripcion", "descripcion del movimiento", "descripcion de la transaccion",
                                "clase de movimiento", "concepto", "detalle", "narracion", "referencia"],
                "monto_unico": ["valor", "monto", "valor original"],
                "cargos": ["cargos", "debitos"],
                "abonos": ["abonos", "creditos"]
            }
            
            def encontrar_columna(df, lista_variantes):
                for variante in lista_variantes:
                    candidatos = [c for c in df.columns if variante in c]
                    if candidatos:
                        return candidatos[0]
                return None
            
            col_fecha_banco = encontrar_columna(df_banco_raw, dicc_columnas["fecha"])
            col_desc_banco = encontrar_columna(df_banco_raw, dicc_columnas["descripcion"])
            col_monto_unico = encontrar_columna(df_banco_raw, dicc_columnas["monto_unico"])
            col_cargos = encontrar_columna(df_banco_raw, dicc_columnas["cargos"])
            col_abonos = encontrar_columna(df_banco_raw, dicc_columnas["abonos"])
            
            df_banco = pd.DataFrame()
            df_banco["fecha"] = convertir_fecha_unificada(df_banco_raw[col_fecha_banco])
            df_banco["descripcion"] = (
                df_banco_raw[col_desc_banco].astype(str).str.strip().str.lower().str.replace(r"\s+", " ", regex=True)
            )
            
            if col_monto_unico:
                df_banco["monto"] = pd.to_numeric(df_banco_raw[col_monto_unico], errors="coerce")
            elif col_cargos and col_abonos:
                cargos = pd.to_numeric(df_banco_raw[col_cargos], errors="coerce").fillna(0)
                abonos = pd.to_numeric(df_banco_raw[col_abonos], errors="coerce").fillna(0)
                df_banco["monto"] = abonos - cargos
            
            df_banco = df_banco.dropna(subset=["monto"])

            # ==========================================================
            # 🔍 CONCILIACIÓN POR MONTOS
            # ==========================================================
            
            contab_temp = df_contable.copy()
            banco_temp = df_banco.copy()
            
            contab_no_banco = []
            
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            total_registros = len(contab_temp)
            for contador, (idx, row) in enumerate(contab_temp.iterrows(), start=1):
                progress = contador / total_registros
                progress_bar.progress(progress)
                status_text.text(f"🔍 Conciliando registros... {contador}/{total_registros}")
                
                monto = row["monto"]
                match = banco_temp[banco_temp["monto"] == monto]
                if not match.empty:
                    banco_temp = banco_temp.drop(match.index[0])
                else:
                    contab_no_banco.append(row)
            
            progress_bar.empty()
            status_text.empty()
            
            banco_no_contab = banco_temp.copy()
            
            contab_no_banco = pd.DataFrame(contab_no_banco)
            if contab_no_banco.empty:
                contab_no_banco = pd.DataFrame(columns=["fecha", "descripcion", "monto"])
            
            if banco_no_contab.empty:
                banco_no_contab = pd.DataFrame(columns=["fecha", "descripcion", "monto"])
            
            # Clasificar partidas
            abonos_contab_no_banco = contab_no_banco[contab_no_banco["monto"] > 0].copy()
            abonos_contab_no_banco["Asignacion"] = "Abonos en contabilidad mas no en extractos (Menos)"
            
            cargos_contab_no_banco = contab_no_banco[contab_no_banco["monto"] < 0].copy()
            cargos_contab_no_banco["Asignacion"] = "Cargos en contabilidad mas no en extractos (Mas)"
            
            abonos_banco_no_contab = banco_no_contab[banco_no_contab["monto"] > 0].copy()
            abonos_banco_no_contab["Asignacion"] = "Abonos en extractos mas no en contabilidad (Mas)"
            
            cargos_banco_no_contab = banco_no_contab[banco_no_contab["monto"] < 0].copy()
            cargos_banco_no_contab["Asignacion"] = "Cargos en extractos mas no en contabilidad (Menos)"

            # ==========================================================
            # 💰 IDENTIFICACIÓN DE INGRESOS/GASTOS BANCARIOS
            # ==========================================================
            
            dicc_bancario = [
                "ABONO INTERESES AHORROS","ABONO INTERESES GANADOS","ABONO X GRAVAMEN MOVIMIENTO FINANCIERO",
                "GRAVAMEN MOVIMIENTO FINANCIERO","GRAVAMEN A LOS MOVIMIENTOS FINANCIEROS","AJUSTE INTERES AHORROS DB",
                "CARGO IVA","COBRO CUOT MANEJO TARJ DEBITO","COBRO DE COMISION POR EL USO DEL PORTAL BUSINESS",
                "COBRO PAGO PROVEEDORES OLROS BANCOS","COBRO PAGO PROVEEDORES OTROS BANCOS","COBRO SERVICIO EMPRESARIAL",
                "COBRO SERVICIO MANEJO PORTAL","COBRO TRANSF. ENVIADA OTRA ENTIDAD","COM.IVA MES B.VIR","COM.MES B.VIRTUAL",
                "COMIS RETIRO CAJERO NO BANCOL","COMISION CHEQUE DEVUELTO","COMISION POR USO CAJERO OTRA ENTIDAD",
                "COMISION TRANSF. ENVIADA OTRA ENTIDAD","COMISION TRANSF. ENVIADA OTRA ENTIDAD B","COMISIONES",
                "COSTO CHEQUERA","COSTO CHEQUERA X 25 CHEQUES","COSTO CHEQUERA X 50 CHEQUES","CUOTA MANEJO TARJETA DEBITO",
                "DEVOLUCION COMISION","GMF - RETIRO SUCURSAL","IMP. GMF","IVA COMISION TRANSF. ENVIADA",
                "IVA COMISION TRANSF. ENVIADA OTRA ENTIDAD","IVA COMISION TRANSF. ENVIADA OTRA ENTIDAD B",
                "IVA COMISIONES","IVA POR SERVICIOS","IVA COSTO CHEQUERA","IVA GMF - RETIRO SUCURSAL",
                "IVA TRANSFERENCIA ENVIADA","IVA TRANSFERENCIA ENVIADA OTRA ENTIDAD",
                "COMISION PAGO PROVEEDORES OTROS BANCOS","RENDIMIENTOS FINANCIEROS",
                "REINTEGRO GRAVAMEN MVTO FINANCIERO", "COBRO CONSULTA SALDOS Y MOVIMIENTOS",
                "DESCUENTO SOLICITUD COPIA EXTRACTO", "IMPTO GOBIERNO 4X1000", "CUOTA MANEJO TRJ DEB",
                "CXC IMPTO GOBIERNO 4X1000 MON", "AJUSTE X GRAVAMEN MOVIMIENTO FINANCIER",
                "COBRO IVA SERVICIOS FINANCIEROS", "COBRO CUOTA DE MANEJO TARJETA DEBITO",
                "RENDIMIENTOS FINANCIEROS.", 
            ]
            dicc_prefix = [" ".join(concept.split()[:2]).lower() for concept in dicc_bancario]
            
            diferencias = pd.concat([
                abonos_contab_no_banco, abonos_banco_no_contab,
                cargos_contab_no_banco, cargos_banco_no_contab
            ], ignore_index=True)
            
            diferencias["grupo_ing_gas"] = diferencias["descripcion"].apply(lambda x: identificar_ing_gas(x, dicc_prefix))
            ing_gas = diferencias[diferencias["grupo_ing_gas"].notna()].copy()
            diferencias = diferencias[~diferencias["descripcion"].isin(ing_gas["descripcion"])]
            
            # Reclasificar después de extraer ingresos/gastos bancarios
            abonos_contab_no_banco = diferencias[diferencias["Asignacion"].str.contains("Abonos en contabilidad", case=False, na=False)]
            abonos_banco_no_contab = diferencias[diferencias["Asignacion"].str.contains("Abonos en extractos", case=False, na=False)]
            cargos_contab_no_banco = diferencias[diferencias["Asignacion"].str.contains("Cargos en contabilidad", case=False, na=False)]
            cargos_banco_no_contab = diferencias[diferencias["Asignacion"].str.contains("Cargos en extractos", case=False, na=False)]
            
            ing_gas_consolidado = (
                ing_gas.groupby("descripcion", as_index=False)["monto"].sum().sort_values("descripcion")
            )
            ing_gas_consolidado["fecha"] = ""

            # ==========================================================
            # 💵 CÁLCULO DE SALDO CONTABLE
            # ==========================================================
            
            col_saldo_mov = buscar_columna(df_contable_raw, "saldo movimiento")
            if col_saldo_mov:
                df_contable_raw[col_saldo_mov] = pd.to_numeric(df_contable_raw[col_saldo_mov], errors="coerce")
                saldo_series = df_contable_raw[col_saldo_mov].dropna()
                if not saldo_series.empty:
                    saldo_contabilidad = saldo_series.iloc[-1]
                    st.info(f"💰 Saldo contable: ${saldo_contabilidad:,.2f}")
                else:
                    saldo_contabilidad = df_contable["monto"].sum()
            else:
                saldo_contabilidad = df_contable["monto"].sum()

            # ==========================================================
            # 📊 MÉTRICAS DE CONCILIACIÓN
            # ==========================================================
            
            st.markdown("---")
            st.subheader("📊 Resumen de Conciliación")
            
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                st.metric(
                    "Abonos Contab. no Banco",
                    f"${abs(abonos_contab_no_banco['monto'].sum()):,.0f}",
                    delta=f"{len(abonos_contab_no_banco)} registros"
                )
            
            with col2:
                st.metric(
                    "Abonos Banco no Contab.",
                    f"${abs(abonos_banco_no_contab['monto'].sum()):,.0f}",
                    delta=f"{len(abonos_banco_no_contab)} registros"
                )
            
            with col3:
                st.metric(
                    "Cargos Contab. no Banco",
                    f"${abs(cargos_contab_no_banco['monto'].sum()):,.0f}",
                    delta=f"{len(cargos_contab_no_banco)} registros"
                )
            
            with col4:
                st.metric(
                    "Ingresos/Gastos Bancarios",
                    f"${abs(ing_gas_consolidado['monto'].sum()):,.0f}",
                    delta=f"{len(ing_gas_consolidado)} conceptos"
                )

            # ==========================================================
            # 📄 GENERACIÓN DEL ARCHIVO EXCEL
            # ==========================================================
            
            st.markdown("---")
            with st.spinner("📝 Generando archivo Excel..."):
                
                wb = Workbook()
                ws = wb.active
                ws.title = "Conciliacion"
                ws.sheet_view.showGridLines = False
                
                vino = "800000"
                mostaza = "FFCC29"
                fill_mostaza = PatternFill(start_color=mostaza, end_color=mostaza, fill_type="solid")
                border_bottom = Border(bottom=Side(style="thin", color=vino))
                
                # ==========================================================
                # Fila 1: Encabezado principal
                # ==========================================================
                ws.merge_cells("A1:C1")
                ws["A1"] = "CONCILIACIÓN BANCARIA"
                ws["A1"].font = Font(name="Calibri", bold=True, size=20, color="000000")
                ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
                
                borde_grueso = Side(style="medium", color=vino)
                borde_doble = Side(style="double", color=vino)
                
                for col in range(1, 4):
                    ws.cell(row=1, column=col).border = Border(
                        top=borde_grueso,
                        left=borde_grueso if col == 1 else Side(style=None),
                        right=borde_grueso if col == 3 else Side(style=None),
                        bottom=borde_doble
                    )
                
                # ==========================================================
                # Fila 2: Nombre empresa
                # ==========================================================
                try:
                    uploaded_contab.seek(0)
                    nombre_empresa = str(pd.read_excel(uploaded_contab, header=None).iloc[2, 0]).strip()
                except:
                    nombre_empresa = "NOMBRE NO DISPONIBLE"
                
                relleno_empresa = "EEEDC0"
                ws.merge_cells("A2:C2")
                ws["A2"] = nombre_empresa
                ws["A2"].font = Font(name="Calibri", size=16, bold=False, color="000000")
                ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
                ws["A2"].fill = PatternFill(start_color=relleno_empresa, end_color=relleno_empresa, fill_type="solid")
                
                for col in range(1, 4):
                    ws.cell(row=2, column=col).border = Border(
                        left=borde_grueso if col == 1 else Side(style=None),
                        right=borde_grueso if col == 3 else Side(style=None),
                        bottom=borde_grueso
                    )
                
                # ==========================================================
                # Filas 3-6: Metadatos
                # ==========================================================
                relleno_vino = PatternFill(start_color=vino, end_color=vino, fill_type="solid")
                
                # Mes
                try:
                    uploaded_contab.seek(0)
                    texto_periodo = str(pd.read_excel(uploaded_contab, header=None).iloc[4, 0]).strip()
                    meses = ["ENERO","FEBRERO","MARZO","ABRIL","MAYO","JUNIO","JULIO","AGOSTO",
                             "SEPTIEMBRE","OCTUBRE","NOVIEMBRE","DICIEMBRE"]
                    mes_encontrado = ""
                    for m in meses:
                        if re.search(m, texto_periodo.upper()):
                            mes_encontrado = m
                    if not mes_encontrado:
                        mes_encontrado = "MES NO DETECTADO"
                except:
                    mes_encontrado = "NO DISPONIBLE"
                
                ws["A3"].fill = relleno_vino
                ws["B3"] = "Mes"
                ws["B3"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
                ws["B3"].alignment = Alignment(horizontal="right", vertical="center")
                ws["B3"].fill = relleno_vino
                ws["C3"] = mes_encontrado
                ws["C3"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
                ws["C3"].alignment = Alignment(horizontal="right", vertical="center")
                ws["C3"].fill = relleno_vino
                
                # Código contable
                try:
                    uploaded_contab.seek(0)
                    codigo_contable = str(pd.read_excel(uploaded_contab, header=None).iloc[9, 0]).strip()
                except:
                    codigo_contable = "NO DISPONIBLE"
                
                ws["A4"].fill = relleno_vino
                ws["B4"] = "Código contable"
                ws["B4"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
                ws["B4"].alignment = Alignment(horizontal="right", vertical="center")
                ws["B4"].fill = relleno_vino
                ws["C4"] = codigo_contable
                ws["C4"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
                ws["C4"].alignment = Alignment(horizontal="right", vertical="center")
                ws["C4"].fill = relleno_vino
                
                # ID Cuenta bancaria
                try:
                    uploaded_contab.seek(0)
                    id_cuenta_bancaria = str(pd.read_excel(uploaded_contab, header=None).iloc[9, 1]).strip()
                except:
                    id_cuenta_bancaria = "NO DISPONIBLE"
                
                ws["A5"].fill = relleno_vino
                ws["B5"] = "ID Cuenta bancaria"
                ws["B5"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
                ws["B5"].alignment = Alignment(horizontal="right", vertical="center")
                ws["B5"].fill = relleno_vino
                ws["C5"] = id_cuenta_bancaria
                ws["C5"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
                ws["C5"].alignment = Alignment(horizontal="right", vertical="center")
                ws["C5"].fill = relleno_vino
                
                # Fecha conciliación
                fecha_actual = datetime.datetime.now().strftime("%d/%m/%Y")
                ws["A6"].fill = relleno_vino
                ws["B6"] = "Fecha de la conciliación"
                ws["B6"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
                ws["B6"].alignment = Alignment(horizontal="right", vertical="center")
                ws["B6"].fill = relleno_vino
                ws["C6"] = fecha_actual
                ws["C6"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
                ws["C6"].alignment = Alignment(horizontal="right", vertical="center")
                ws["C6"].fill = relleno_vino
                
                # ==========================================================
                # Cuadro de verificación
                # ==========================================================
                ws.append(["", "Concepto", "Monto"])
                for cell in ws[ws.max_row]:
                    cell.font = Font(bold=True, size=14)
                    cell.fill = fill_mostaza
                    cell.border = border_bottom
                
                ws.append(["", "SALDO CONTABILIDAD", saldo_contabilidad])
                fila_saldo_contab = ws.max_row
                for cell in ws[ws.max_row]:
                    cell.font = Font(bold=True, size=14)
                    cell.fill = fill_mostaza
                    cell.border = border_bottom
                
                ws.append(["", "SALDO PARTIDAS", ""])
                fila_saldo_partidas = ws.max_row
                for cell in ws[ws.max_row]:
                    cell.font = Font(bold=True, size=14)
                    cell.fill = fill_mostaza
                    cell.border = border_bottom
                
                ws.append(["", "CIFRA VERIFICACION", ""])
                fila_cifra_verif = ws.max_row
                for cell in ws[ws.max_row]:
                    cell.font = Font(bold=True, size=14)
                    cell.fill = fill_mostaza
                    cell.border = border_bottom
                
                ws.append(["", "SALDO EXTRACTO", ""])
                fila_saldo_extracto = ws.max_row
                for cell in ws[ws.max_row]:
                    cell.font = Font(bold=True, size=14)
                    cell.fill = fill_mostaza
                    cell.border = border_bottom
                
                ws.append(["", "DIFERENCIA", ""])
                fila_diferencia = ws.max_row
                for cell in ws[ws.max_row]:
                    cell.font = Font(bold=True, size=14)
                    cell.fill = fill_mostaza
                    cell.border = border_bottom
                
                # ==========================================================
                # Función para escribir bloques con fórmulas
                # ==========================================================
                dicc_titulos = {
                    "abonos_contab_no_banco": "ABONOS EN CONTABILIDAD MAS NO EN EXTRACTOS (MENOS)",
                    "abonos_banco_no_contab": "ABONOS EN EXTRACTOS MAS NO EN CONTABILIDAD (MAS)",
                    "cargos_contab_no_banco": "CARGOS EN CONTABILIDAD MAS NO EN EXTRACTOS (MAS)",
                    "cargos_banco_no_contab": "CARGOS EN EXTRACTOS MAS NO EN CONTABILIDAD (MENOS)",
                    "ing_gas_consolidado": "INGRESOS / GASTOS BANCARIOS (MENOS)"
                }
                
                bloques_info = []
                filas_referencias = {}
                
                def escribir_bloque(df, titulo):
                    ws.append([""])
                    ws.append([titulo, "", ""])
                    fila_subtotal = ws.max_row
                    
                    for cell in ws[ws.max_row]:
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color=vino, end_color=vino, fill_type="solid")
                        cell.alignment = Alignment(horizontal="left")
                    
                    ws.append(["Fecha", "Descripción", "Monto"])
                    for cell in ws[ws.max_row]:
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal="center")
                    
                    fila_inicio_datos = ws.max_row + 1
                    
                    if df.empty:
                        ws.append(["", "(Sin partidas registradas)", 0])
                    else:
                        for _, row in df.iterrows():
                            fecha = "" if "fecha" not in row or pd.isna(row["fecha"]) else (
                                row["fecha"].strftime("%Y-%m-%d") if isinstance(row["fecha"], (datetime.datetime, pd.Timestamp)) else row["fecha"]
                            )
                            ws.append([fecha, row["descripcion"], row["monto"]])
                    
                    fila_fin_datos = ws.max_row
                    
                    bloques_info.append({
                        'fila_subtotal': fila_subtotal,
                        'fila_inicio': fila_inicio_datos,
                        'fila_fin': fila_fin_datos
                    })
                    
                    return fila_subtotal
                
                # Escribir todos los bloques
                filas_referencias["abonos_contab"] = escribir_bloque(abonos_contab_no_banco, dicc_titulos["abonos_contab_no_banco"])
                filas_referencias["abonos_banco"] = escribir_bloque(abonos_banco_no_contab, dicc_titulos["abonos_banco_no_contab"])
                filas_referencias["cargos_contab"] = escribir_bloque(cargos_contab_no_banco, dicc_titulos["cargos_contab_no_banco"])
                filas_referencias["cargos_banco"] = escribir_bloque(cargos_banco_no_contab, dicc_titulos["cargos_banco_no_contab"])
                filas_referencias["ing_gas"] = escribir_bloque(ing_gas_consolidado, dicc_titulos["ing_gas_consolidado"])
                
                # ==========================================================
                # Actualizar fórmulas de subtotales
                # ==========================================================
                for i, bloque in enumerate(bloques_info):
                    fila_subtotal = bloque['fila_subtotal']
                    fila_inicio = bloque['fila_inicio']
                    
                    if i < len(bloques_info) - 1:
                        fila_inicio_siguiente = bloques_info[i + 1]['fila_inicio']
                        fila_fin = fila_inicio_siguiente - 3
                    else:
                        fila_fin = bloque['fila_fin'] + 1
                    
                    if fila_fin >= fila_inicio:
                        formula_subtotal = f"=ABS(SUM(C{fila_inicio}:C{fila_fin}))"
                    else:
                        formula_subtotal = "=0"
                    
                    ws[f"C{fila_subtotal}"] = formula_subtotal
                
                # ==========================================================
                # Insertar fórmulas en el cuadro de verificación
                # ==========================================================
                formula_saldo_partidas = (
                    f"=-C{filas_referencias['abonos_contab']}+C{filas_referencias['abonos_banco']}"
                    f"+C{filas_referencias['cargos_contab']}-C{filas_referencias['cargos_banco']}"
                    f"-C{filas_referencias['ing_gas']}"
                )
                ws[f"C{fila_saldo_partidas}"] = formula_saldo_partidas
                
                ws[f"C{fila_cifra_verif}"] = f"=C{fila_saldo_contab}+C{fila_saldo_partidas}"
                ws[f"C{fila_diferencia}"] = f"=C{fila_cifra_verif}-C{fila_saldo_extracto}"
                
                # ==========================================================
                # Formato de números y alineación
                # ==========================================================
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=3, max_col=3):
                    for cell in row:
                        if cell.value or cell.data_type == 'f':
                            cell.number_format = u'_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'
                
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
                    for cell in row:
                        cell.alignment = Alignment(horizontal="center")
                
                for i, width in enumerate([15, 80, 20], start=1):
                    ws.column_dimensions[get_column_letter(i)].width = width
                
                titulos_set = set(dicc_titulos.values())
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
                    for cell in row:
                        if cell.value in titulos_set:
                            cell.alignment = Alignment(horizontal="left")
                
                # ==========================================================
                # Guardar archivo en memoria
                # ==========================================================
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                
                st.success("✅ Conciliación generada correctamente")
            
            # ==========================================================
            # 📥 DESCARGA DEL ARCHIVO
            # ==========================================================
            st.markdown("---")
            st.subheader("📥 Descargar Conciliación")
            
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            nombre_archivo = f"Conciliacion_Bancaria_{timestamp}.xlsx"
            
            st.download_button(
                label="📄 Descargar Conciliación Bancaria",
                data=output,
                file_name=nombre_archivo,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="primary"
            )
            
            # ==========================================================
            # 📊 VISTA PREVIA DE PARTIDAS
            # ==========================================================
            with st.expander("📋 Ver Detalle de Partidas Conciliadas"):
                tab1, tab2, tab3, tab4, tab5 = st.tabs([
                    "Abonos Contab. no Banco",
                    "Abonos Banco no Contab.",
                    "Cargos Contab. no Banco",
                    "Cargos Banco no Contab.",
                    "Ingresos/Gastos Bancarios"
                ])
                
                with tab1:
                    if not abonos_contab_no_banco.empty:
                        st.dataframe(
                            abonos_contab_no_banco[["fecha", "descripcion", "monto"]],
                            use_container_width=True
                        )
                    else:
                        st.info("✅ No hay partidas en esta categoría")
                
                with tab2:
                    if not abonos_banco_no_contab.empty:
                        st.dataframe(
                            abonos_banco_no_contab[["fecha", "descripcion", "monto"]],
                            use_container_width=True
                        )
                    else:
                        st.info("✅ No hay partidas en esta categoría")
                
                with tab3:
                    if not cargos_contab_no_banco.empty:
                        st.dataframe(
                            cargos_contab_no_banco[["fecha", "descripcion", "monto"]],
                            use_container_width=True
                        )
                    else:
                        st.info("✅ No hay partidas en esta categoría")
                
                with tab4:
                    if not cargos_banco_no_contab.empty:
                        st.dataframe(
                            cargos_banco_no_contab[["fecha", "descripcion", "monto"]],
                            use_container_width=True
                        )
                    else:
                        st.info("✅ No hay partidas en esta categoría")
                
                with tab5:
                    if not ing_gas_consolidado.empty:
                        st.dataframe(
                            ing_gas_consolidado[["descripcion", "monto"]],
                            use_container_width=True
                        )
                    else:
                        st.info("✅ No hay partidas en esta categoría")
        
        except Exception as e:
            st.error(f"❌ Error durante el procesamiento: {str(e)}")
            st.exception(e)

# ==========================================================
# 📖 INFORMACIÓN ADICIONAL
# ==========================================================
with st.expander("📘 Instrucciones de Uso"):
    st.markdown("""
    ### 🏦 Cómo usar el Generador de Conciliación Bancaria:
    
    1. **📊 Archivo de Contabilidad:**
       - Carga el archivo Excel del reporte “Movimiento auxiliar por cuenta contable” 
         tal como se descarga del sistema, sin aplicar ninguna edición.
       - Debe incluir columnas: Fecha, Comprobante, Tercero, Débito, Crédito.
       - El sistema detecta automáticamente las columnas relevantes.
    
    <br>
    
    2. **🏦 Extracto Bancario:**
       - Sube el extracto bancario en formato excel.
       - Debe contener: Fecha, Descripción y Monto (Cargos/Abonos en una misma columna).
    Nota: Asegurate de que el archivo no contenga columnas vacias o innecesarias.
    
    <br>
    
    3. **⚙️ Procesamiento Automático:**
       - El sistema concilia automáticamente los montos coincidentes.
       - Identifica partidas pendientes en ambos archivos.
       - Clasifica ingresos y gastos bancarios automáticamente.
    
    <br>
    
    4. **📥 Descarga y Revisión:**
       - Descarga el archivo Excel generado con todas las fórmulas activas.
       - Completa el campo **"SALDO EXTRACTO"** en la fila correspondiente.
       - La **DIFERENCIA** se calculará automáticamente.
    
    <br>
    
    5. **💡 Consejos:**
       - Verifica que las fechas estén en formato "DD/MM/YYYY".
       - Revisa las partidas clasificadas en el expandible de "Detalle".
       - Si hay diferencias, revisa los conceptos bancarios no identificados.
       - Si asigna un concepto de ingreso o gasto bancario en una sección distinta
         a la indicada, informe al área de Business Intelligence.
    """, unsafe_allow_html=True)

# ==========================================================
# 🧩 CARACTERÍSTICAS DE LA CONCILIACIÓN
# ==========================================================
with st.expander("🧩 Características de la Conciliación Bancaria"):
    st.markdown("""
    ### ✨ Funcionalidades incluidas:
    
    - **🔍 Conciliación automática por montos**: Cruza los movimientos contables con el extracto bancario.
    - **📊 Clasificación inteligente**: Separa automáticamente abonos, cargos e ingresos/gastos bancarios.
    - **🧮 Fórmulas dinámicas**: El archivo Excel incluye fórmulas activas para cálculos automáticos.
    - **🎨 Formato profesional**: Diseño visual con colores institucionales y bordes personalizados.
    - **📅 Detección automática de fechas**: Soporta múltiples formatos de fecha (DD/MM/YYYY, YYYY-MM-DD, etc.).
    - **💰 Identificación de gastos bancarios**: Reconoce automáticamente más de 40 conceptos bancarios comunes.
    - **📋 Metadata automática**: Extrae información del encabezado del archivo contable (empresa, código, cuenta).
    - **🔄 Rangos dinámicos en fórmulas**: Los subtotales se ajustan automáticamente al contenido.
    - **✅ Sin cuadrículas**: Vista limpia y profesional del documento final.
    """)
